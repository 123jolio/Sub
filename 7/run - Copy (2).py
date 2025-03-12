import os
import re
import logging
import time
import numpy as np
import pandas as pd
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from multiprocessing import Pool, cpu_count
import matplotlib
matplotlib.use("TkAgg")  # Force interactive TkAgg backend
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button, RadioButtons
from tkinter import Tk, simpledialog, messagebox
import win32com.client as win32
import xml.etree.ElementTree as ET  # for parsing KML

# Additional imports for mapping and polygon point-inclusion
import contextily as ctx
from pyproj import Transformer
from matplotlib.path import Path

logging.basicConfig(level=logging.INFO, format=r'%(asctime)s - %(levelname)s - %(message)s')
working_dir = os.getcwd()

# ============================================================================
# Helper Functions
# ============================================================================

def find_file(folder, pattern):
    """Return the first file in folder matching the regex pattern (case-insensitive)."""
    regex = re.compile(pattern, re.IGNORECASE)
    for fname in os.listdir(folder):
        if regex.search(fname):
            return os.path.join(folder, fname)
    return None

def parse_kml_bounds(kml_path):
    """Extract (lon_min, lon_max, lat_min, lat_max) from a KML file."""
    try:
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        coords_text = root.find('.//kml:coordinates', ns).text.strip()
        coords = coords_text.split()
        unique_coords = []
        for coord in coords:
            parts = coord.split(',')
            if len(parts) >= 2:
                lon, lat = float(parts[0]), float(parts[1])
                if (lon, lat) not in unique_coords:
                    unique_coords.append((lon, lat))
        if len(unique_coords) < 4:
            return None
        lon_min = unique_coords[0][0]
        lat_max = unique_coords[0][1]
        lon_max = unique_coords[1][0]
        lat_min = unique_coords[3][1]
        return lon_min, lon_max, lat_min, lat_max
    except Exception as e:
        logging.error(f"Error parsing KML file {kml_path}: {e}")
        return None

def set_calculation_properties(workbook):
    """Force auto calculation in openpyxl (if supported)."""
    try:
        cp = workbook.calculation_properties
        cp.calculationMode = "auto"
        cp.calcCompleted = False
        cp.calcOnSave = True
        cp.fullCalcOnLoad = True
        cp.forceFullCalc = True
    except AttributeError:
        logging.warning("Calculation properties not supported; consider upgrading openpyxl.")

def force_recalc_with_excel(final_file):
    """Force full recalculation in Excel via COM Automation."""
    if not os.path.isfile(final_file):
        logging.error(f"Cannot recalc: File not found: {final_file}")
        return
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(final_file, UpdateLinks=False)
        wb.Application.CalculateFullRebuild()
        while wb.Application.CalculationState != 0:
            time.sleep(0.5)
        wb.Save()
        wb.Close(False)
        excel.Quit()
        logging.info(f"Excel recalculation forced and file saved: {final_file}")
    except Exception as e:
        logging.error(f"Error during Excel recalculation: {e}")

def is_indicator_complete(result_folder, indicator):
    """Return True if a chart file matching 'chart_{indicator}.xlsx' exists in result_folder."""
    pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
    return bool(find_file(result_folder, pattern))

# ============================================================================
# Interactive Grid Creation on Map
# ============================================================================

def create_grid_on_map(indicator, jpgs_dir):
    """
    Provides two options for grid creation:
      1. Polygon Mode: The user clicks multiple times to define a polygon.
         Then the user is asked to input an initial grid spacing.
         A grid is generated over the (rotated) polygon's bounding box and only points inside the polygon are kept.
         In the preview window, two sliders adjust horizontal and vertical spacing (density), one slider rotates the grid,
         and one slider rotates the polygon (i.e. adjusts the angles that define it).
      2. Origin Mode: (Fallback) The user clicks one point as the origin and uses sliders
         (for spacing, rows, cols, and angle) to create a grid.
    In both cases, grid points are sorted (left-to-right, bottom-to-top) and saved as "grid_{indicator}.xlsx".
    The grid shape (vertical count) is stored in cells D2 (rows) and E2 (cols).
    """
    root = Tk()
    root.withdraw()  # hide main window

    use_polygon = messagebox.askyesno("Grid Creation Mode",
        "Do you want to define the grid by drawing a polygon?\n(If No, you'll set an origin and adjust parameters with sliders.)",
        parent=root)
    
    grid_pts = []
    if use_polygon:
        # User defines polygon vertices.
        messagebox.showinfo("Polygon Input", "Click on the image to define the polygon vertices. Press Enter when done.", parent=root)
        fig_poly, ax_poly = plt.subplots()
        candidates = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
        if not candidates:
            messagebox.showerror("Error", "No background image found in " + jpgs_dir, parent=root)
            root.destroy()
            return None
        bg_path = os.path.join(jpgs_dir, "frame_0.jpg") if os.path.isfile(os.path.join(jpgs_dir, "frame_0.jpg")) else os.path.join(jpgs_dir, candidates[0])
        bg_img = Image.open(bg_path)
        ax_poly.imshow(bg_img, extent=[0, 512, 0, 512])
        ax_poly.set_title("Draw polygon vertices then press Enter")
        original_polygon = plt.ginput(n=-1, timeout=0)
        plt.close(fig_poly)
        if len(original_polygon) < 3:
            messagebox.showerror("Error", "A polygon must have at least 3 points.", parent=root)
            root.destroy()
            return None

        # Compute polygon centroid.
        poly_arr = np.array(original_polygon)
        centroid = poly_arr.mean(axis=0)

        # Function to rotate a set of points about a center.
        def rotate_points(points, angle_deg, center):
            angle_rad = np.radians(angle_deg)
            cos_a, sin_a = np.cos(angle_rad), np.sin(angle_rad)
            rotated = []
            for pt in points:
                shifted = np.array(pt) - center
                rotated_pt = np.array([shifted[0]*cos_a - shifted[1]*sin_a,
                                        shifted[0]*sin_a + shifted[1]*cos_a]) + center
                rotated.append(rotated_pt.tolist())
            return rotated

        # Set initial values.
        spacing_h_init = float(simpledialog.askstring("Grid Spacing", "Enter initial horizontal grid spacing:", parent=root))
        spacing_v_init = spacing_h_init  # initial vertical spacing same as horizontal
        grid_rot_init = 0.0
        poly_rot_init = 0.0

        # Function to generate grid points given parameters.
        def generate_grid(h_spacing, v_spacing, poly_rot, grid_rot):
            # Rotate polygon by poly_rot.
            adjusted_poly = rotate_points(original_polygon, poly_rot, centroid)
            poly_path = Path(adjusted_poly)
            # Compute bounding box of adjusted polygon.
            xs, ys = zip(*adjusted_poly)
            bx_min, bx_max = min(xs), max(xs)
            by_min, by_max = min(ys), max(ys)
            # Generate grid over bounding box.
            pts = []
            x_coords = np.arange(bx_min, bx_max + h_spacing, h_spacing)
            y_coords = np.arange(by_min, by_max + v_spacing, v_spacing)
            for y in y_coords:
                for x in x_coords:
                    pts.append([x, y])
            # Rotate the grid about the center of the bounding box.
            grid_center = [ (bx_min+bx_max)/2, (by_min+by_max)/2 ]
            pts_rot = rotate_points(pts, grid_rot, grid_center)
            # Keep only points inside the adjusted polygon.
            valid_pts = [pt for pt in pts_rot if poly_path.contains_point((pt[0], pt[1]))]
            valid_pts.sort(key=lambda pt: (pt[1], pt[0]))
            return adjusted_poly, valid_pts

        # Create preview figure.
        fig_preview, ax_preview = plt.subplots()
        plt.subplots_adjust(bottom=0.35)
        ax_preview.imshow(bg_img, extent=[0, 512, 0, 512])
        preview_scatter = ax_preview.scatter([], [], color='red', s=10, label='Grid Points')
        poly_line, = ax_preview.plot([], [], color='blue', linestyle='--', linewidth=2, label='Polygon')
        ax_preview.set_title("Preview Grid.\nAdjust sliders then close window when satisfied.")
        ax_preview.legend()

        # Slider axes.
        ax_hspacing = plt.axes([0.15, 0.25, 0.7, 0.03])
        ax_vspacing = plt.axes([0.15, 0.20, 0.7, 0.03])
        ax_gridrot = plt.axes([0.15, 0.15, 0.7, 0.03])
        ax_polyrot = plt.axes([0.15, 0.10, 0.7, 0.03])

        slider_hspacing = Slider(ax_hspacing, 'Horiz Spacing', 1, 100, valinit=spacing_h_init)
        slider_vspacing = Slider(ax_vspacing, 'Vert Spacing', 1, 100, valinit=spacing_v_init)
        slider_gridrot = Slider(ax_gridrot, 'Grid Rotation', -180, 180, valinit=grid_rot_init)
        slider_polyrot = Slider(ax_polyrot, 'Polygon Rotation', -180, 180, valinit=poly_rot_init)

        def update(val):
            h_spacing = slider_hspacing.val
            v_spacing = slider_vspacing.val
            grid_rot = slider_gridrot.val
            poly_rot = slider_polyrot.val
            adjusted_poly, new_grid = generate_grid(h_spacing, v_spacing, poly_rot, grid_rot)
            if new_grid:
                preview_scatter.set_offsets(np.array(new_grid))
            else:
                preview_scatter.set_offsets([])
            # Update polygon outline.
            poly_arr = np.array(adjusted_poly)
            poly_line.set_data(np.append(poly_arr[:,0], poly_arr[0,0]), np.append(poly_arr[:,1], poly_arr[0,1]))
            fig_preview.canvas.draw_idle()

        slider_hspacing.on_changed(update)
        slider_vspacing.on_changed(update)
        slider_gridrot.on_changed(update)
        slider_polyrot.on_changed(update)

        # Initial update.
        update(None)
        plt.show()  # Wait until the user closes the preview window.
        # After preview, use the last computed grid.
        _, grid_pts = generate_grid(slider_hspacing.val, slider_vspacing.val, slider_polyrot.val, slider_gridrot.val)

    else:
        # Origin Mode (fallback): Let the user click one point for origin and use sliders.
        logging.info("Using Origin Mode for grid creation.")
        fig_origin, ax_origin = plt.subplots()
        candidates = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
        if not candidates:
            messagebox.showerror("Error", "No background image found in " + jpgs_dir, parent=root)
            root.destroy()
            return None
        bg_path = os.path.join(jpgs_dir, "frame_0.jpg") if os.path.isfile(os.path.join(jpgs_dir, "frame_0.jpg")) else os.path.join(jpgs_dir, candidates[0])
        bg_img = Image.open(bg_path)
        ax_origin.imshow(bg_img, extent=[0, 512, 0, 512])
        ax_origin.set_title("Click to set grid origin")
        origin = plt.ginput(1)
        if not origin:
            plt.close(fig_origin)
            root.destroy()
            return None
        origin = origin[0]
        logging.info(f"User selected origin: {origin}")
        plt.close(fig_origin)

        fig, ax = plt.subplots()
        ax.imshow(bg_img, extent=[0, 512, 0, 512])
        ax.set_title("Adjust grid parameters and click 'Save Grid'")
        init_spacing = 50
        init_rows = 10
        init_cols = 10
        init_angle = 0

        slider_ax_angle = plt.axes([0.25, 0.20, 0.65, 0.03])
        slider_angle = Slider(slider_ax_angle, 'Angle', -90, 90, valinit=init_angle)
        slider_ax_spacing = plt.axes([0.25, 0.15, 0.65, 0.03])
        slider_spacing = Slider(slider_ax_spacing, 'Spacing', 10, 200, valinit=init_spacing)
        slider_ax_rows = plt.axes([0.25, 0.10, 0.65, 0.03])
        slider_rows = Slider(slider_ax_rows, 'Rows', 1, 50, valinit=init_rows, valfmt='%0.0f')
        slider_ax_cols = plt.axes([0.25, 0.05, 0.65, 0.03])
        slider_cols = Slider(slider_ax_cols, 'Cols', 1, 50, valinit=init_cols, valfmt='%0.0f')

        grid_pts = []

        def update_grid(val):
            nonlocal grid_pts
            spacing = slider_spacing.val
            n_rows = int(slider_rows.val)
            n_cols = int(slider_cols.val)
            angle = slider_angle.val
            import math
            angle_rad = math.radians(angle)
            grid_pts = []
            for i in range(n_rows):
                for j in range(n_cols):
                    x = j * spacing
                    y = i * spacing
                    x_rot = x * math.cos(angle_rad) - y * math.sin(angle_rad)
                    y_rot = x * math.sin(angle_rad) + y * math.cos(angle_rad)
                    grid_pts.append([origin[0] + x_rot, origin[1] + y_rot])
            grid_pts.sort(key=lambda pt: (pt[1], pt[0]))
            for coll in ax.collections[:]:
                coll.remove()
            if grid_pts:
                xs, ys = zip(*grid_pts)
                ax.scatter(xs, ys, c='red', s=10)
            fig.canvas.draw_idle()

        slider_spacing.on_changed(update_grid)
        slider_rows.on_changed(update_grid)
        slider_cols.on_changed(update_grid)
        slider_angle.on_changed(update_grid)
        update_grid(None)
        ax_button = plt.axes([0.8, 0.9, 0.1, 0.05])
        button = Button(ax_button, 'Save Grid')
        grid_pts_final = None

        def save_grid(event):
            nonlocal grid_pts_final
            grid_pts_final = grid_pts
            plt.close(fig)

        button.on_clicked(save_grid)
        plt.show()
        if not grid_pts_final:
            root.destroy()
            return None
        grid_pts = grid_pts_final

    root.destroy()  # Clean up the Tkinter root
    if not grid_pts:
        return None

    # Sort points (left-to-right, bottom-to-top)
    grid_pts.sort(key=lambda pt: (pt[1], pt[0]))

    wb = Workbook()
    ws = wb.active
    ws.title = "GridPoints"
    ws.append(["X", "Y"])
    for pt in grid_pts:
        ws.append(pt)
    # For chart processing, record grid shape.
    if use_polygon:
        ys = [pt[1] for pt in grid_pts]
        unique_y = np.unique(np.round(ys, 2))
        n_rows = len(unique_y)
        n_cols = int(np.ceil(len(grid_pts) / n_rows))
    else:
        n_rows = int(slider_rows.val)
        n_cols = int(slider_cols.val)
    ws['D1'] = "Rows"
    ws['D2'] = n_rows
    ws['E1'] = "Cols"
    ws['E2'] = n_cols

    grid_filename = f"grid_{indicator}.xlsx"
    save_path = os.path.join(working_dir, grid_filename)
    wb.save(save_path)
    logging.info(f"Grid file created: {save_path}")
    return save_path

# ============================================================================
# PART A: Grid and Image Processing
# ============================================================================

def read_points_from_xlsx_openpyxl(file_path):
    """Read grid points from the first sheet of a .xlsx file and return (points, grid_shape)."""
    points = []
    grid_shape = (None, None)
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        points = [(float(row[0]), float(row[1])) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None and row[1] is not None]
        n_rows = sheet['D2'].value
        n_cols = sheet['E2'].value
        grid_shape = (int(n_rows) if n_rows is not None else None, int(n_cols) if n_cols is not None else None)
    except Exception as e:
        logging.error(f"Error reading grid file {file_path}: {e}")
    finally:
        wb.close()
    return points, grid_shape

def read_rgb_values_at_point(img, point):
    x, y = point
    row = img.height - int(y) - 1
    col = int(x)
    if 0 <= row < img.height and 0 <= col < img.width:
        return img.getpixel((col, row))
    return None

def process_image(image_path, points):
    img = Image.open(image_path).convert("RGB")
    img_width, img_height = img.size
    scaled_points = [(x * img_width / 512, y * img_height / 512) for x, y in points]
    results = []
    for point in scaled_points:
        rgb = read_rgb_values_at_point(img, point)
        if rgb:
            r, g, b = rgb
            grayscale = 0.299*r + 0.587*g + 0.114*b
            results.append([point[0], point[1],
                            round(r,3), round(g,3), round(b,3),
                            round(grayscale,3)])
    img.close()
    return results

def write_results_to_excel(results, workbook_path):
    wb = Workbook()
    wb.remove(wb.active)
    for index, data in enumerate(results):
        sheet = wb.create_sheet(title=str(index))
        sheet.append(['X', 'Y', 'R', 'G', 'B', 'Grayscale'])
        for row in data:
            sheet.append(row)
    set_calculation_properties(wb)
    wb.save(workbook_path)

def run_partA(input_grid_folder, input_images_folder, output_folder, indicator):
    pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
    grid_file = find_file(input_grid_folder, pattern)
    if not grid_file:
        logging.error(f"Grid file not found in {input_grid_folder} for indicator {indicator}")
        return False
    points, _ = read_points_from_xlsx_openpyxl(grid_file)
    if not points:
        logging.error(f"No points read from {grid_file}. Exiting Part A.")
        return False
    jpg_files = [os.path.join(input_images_folder, f) for f in os.listdir(input_images_folder) if f.lower().endswith('.jpg')]
    if not jpg_files:
        logging.error(f"No .jpg files found in {input_images_folder}.")
        return False
    with Pool(cpu_count()) as pool:
        results = pool.starmap(process_image, [(jpg, points) for jpg in jpg_files])
    output_xlsx = os.path.join(output_folder, f"results_{indicator}.xlsx")
    write_results_to_excel(results, output_xlsx)
    logging.info(f"Part A completed. Wrote {output_xlsx}")
    return True

# ============================================================================
# PART C: Process Total Results
# ============================================================================

def run_partC(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("~$") or filename.startswith("chart_"):
            continue
        if filename.endswith(".xlsx"):
            input_file = os.path.join(output_folder, filename)
            match = re.search(r'_([A-Za-z0-9]+)\.xlsx$', filename)
            if not match:
                continue
            file_indicator = match.group(1)
            out_file = os.path.join(output_folder, f"total_results_{file_indicator}.xlsx")
            wb = load_workbook(input_file, read_only=True, data_only=True)
            data_frames = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = list(sheet.values)
                if not data:
                    continue
                data = data[1:]
                df = pd.DataFrame(data)
                try:
                    col_data = df.iloc[:, 3].dropna()
                    if not col_data.empty:
                        data_frames.append(col_data.reset_index(drop=True).to_frame(name=sheet_name))
                except IndexError:
                    logging.info(f"No Column 'D' in {sheet_name}")
            if data_frames:
                all_data_df = pd.concat(data_frames, axis=1)
                all_data_df['Average'] = all_data_df.mean(axis=1)
                with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                    all_data_df.to_excel(writer, sheet_name='ChlA', index=False)
                    set_calculation_properties(writer.book)
                logging.info(f"Part C: Wrote {out_file}")
            else:
                logging.info(f"No data frames to concatenate for {filename}")

# ============================================================================
# PART D: Process Charts
# ============================================================================

def run_partD(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("~$"):
            continue
        if filename.startswith("total_results_") and filename.endswith(".xlsx"):
            source_file = os.path.join(output_folder, filename)
            file_indicator = indicator  
            wb = load_workbook(source_file, data_only=True)
            if 'ChlA' not in wb.sheetnames:
                logging.info(f"Sheet 'ChlA' not found in {filename}")
                continue
            ws = wb['ChlA']
            avg_col = None
            for col in range(1, ws.max_column+1):
                if ws.cell(row=1, column=col).value == 'Average':
                    avg_col = col
                    break
            if avg_col is None:
                logging.info("Column 'Average' not found.")
                continue
            avg_data = [ws.cell(row=r, column=avg_col).value for r in range(2, ws.max_row+1)]
            grid_file = find_file(os.path.join(os.path.dirname(source_file).replace("_result", "_grid")), rf"grid_.*{re.escape(indicator)}.*\.xlsx")
            _, grid_shape = read_points_from_xlsx_openpyxl(grid_file) if grid_file else (None, (20,))
            grid_rows = grid_shape[0] if grid_shape[0] is not None else 20

            wb_out = Workbook()
            ws1 = wb_out.active
            ws1.title = 'Sheet1'
            for r in range(1, grid_rows+1):
                ws1.cell(row=r, column=2, value=r).number_format = "0.0000"
            col_index = 3
            for i in range(0, len(avg_data), grid_rows):
                for j in range(grid_rows):
                    if i+j < len(avg_data):
                        ws1.cell(row=j+1, column=col_index, value=avg_data[i+j]).number_format = "0.0000"
                col_index += 1
            ws2 = wb_out.create_sheet(title='Sheet2')
            for col in range(3, col_index):
                col_letter = get_column_letter(col)
                for row in range(1, grid_rows+1):
                    ws2.cell(row=row, column=col, value=f"=Sheet1!{col_letter}{row}/AVERAGE(Sheet1!{col_letter}$1:{col_letter}$" + f"{grid_rows})")
            dummy = wb_out.create_sheet(title='DummyCalc')
            dummy['A1'] = "=NOW()"
            dummy.sheet_state = 'hidden'
            set_calculation_properties(wb_out)
            intermediate_file = os.path.join(output_folder, f"chart_{file_indicator}_intermediate.xlsx")
            wb_out.save(intermediate_file)
            wb2 = load_workbook(intermediate_file)
            if 'Sheet2' in wb2.sheetnames:
                ws2_2 = wb2['Sheet2']
                ws3 = wb2.create_sheet(title='Sheet3')
                dest_row = 1
                for col in range(1, ws2_2.max_column+1):
                    for row in range(1, ws2_2.max_row+1):
                        val = ws2_2.cell(row=row, column=col).value
                        if val is not None:
                            ws3.cell(row=dest_row, column=2, value=val)
                            dest_row += 1
                set_calculation_properties(wb2)
                final_chart = os.path.join(output_folder, f"chart_{file_indicator}.xlsx")
                wb2.save(final_chart)
                logging.info(f"Part D: Chart file generated: {final_chart}")

# ============================================================================
# PART E: Export XLSM and CSV (Removed)
# ============================================================================

def run_partE(output_folder, indicator):
    logging.info(f"Part E: Skipping export of XLSM and CSV for indicator {indicator}")

# ============================================================================
# Reading Data for Scatter Plots
# ============================================================================

def read_points_from_xlsx_openpyxl(file_path):
    """Read grid points from the first sheet of a .xlsx file and return (points, grid_shape)."""
    points = []
    grid_shape = (None, None)
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        points = [(float(row[0]), float(row[1])) for row in sheet.iter_rows(min_row=2, values_only=True) if row[0] is not None and row[1] is not None]
        n_rows = sheet['D2'].value
        n_cols = sheet['E2'].value
        grid_shape = (int(n_rows) if n_rows is not None else None, int(n_cols) if n_cols is not None else None)
    except Exception as e:
        logging.error(f"Error reading grid file {file_path}: {e}")
    finally:
        wb.close()
    return points, grid_shape

def read_data_from_xlsx(file_path, sheet_name, column_index):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
    ws = wb[sheet_name]
    return [float(row[0]) for row in ws.iter_rows(min_row=1, min_col=column_index+1, max_col=column_index+1, values_only=True) if row[0] is not None]

# ============================================================================
# Interactive Scatter UI with Basemap
# ============================================================================

def run_interactive_scatter_for_all_indicators(all_data):
    """
    all_data: dict of indicator -> {
         'points': (pixel x_values, pixel y_values),
         'geo_points': (geo x_values, geo y_values) [optional],
         'geo_bounds': [lon_min, lon_max, lat_min, lat_max] [optional],
         'data': data array,
         'background': path to background image
      }
    If geo_points and geo_bounds are provided, they are transformed to Web Mercator
    and an OpenStreetMap basemap is added.
    """
    if not all_data:
        logging.error("No indicator data available for interactive scatter. Exiting.")
        return

    import matplotlib.pyplot as plt
    from PIL import Image

    indicators = list(all_data.keys())
    current_indicator = indicators[0]

    def get_current_data():
        return all_data[current_indicator]

    def update_scatter():
        d = get_current_data()
        if d.get('geo_points') is not None:
            xvals, yvals = d['geo_points']
            transformer = Transformer.from_crs("EPSG:4326", "EPSG:3857", always_xy=True)
            xvals_merc = [transformer.transform(lon, lat)[0] for lon, lat in zip(xvals, d['geo_points'][1])]
            yvals_merc = [transformer.transform(lon, lat)[1] for lon, lat in zip(xvals, d['geo_points'][1])]
            xvals, yvals = xvals_merc, yvals_merc
            title_suffix = " (Geo)"
        else:
            xvals, yvals = d['points']
            title_suffix = " (Pixels)"
        arr = d['data']
        img_path = d['background']
        new_img = Image.open(img_path)
        for im in ax1.images + ax2.images:
            im.set_data(new_img)
        new_sizes = np.full(len(arr), slider_size.val)
        new_colors = []
        for val in arr:
            dist = abs(val - slider_min.val) / max(slider_max.val - slider_min.val, 1e-6)
            gray = int(255 * np.exp(slider_exp.val * dist) * slider_sharp.val)
            gray = max(min(gray, 255), 0)
            new_colors.append((gray/255, gray/255, gray/255))
        scatter1.set_offsets(np.column_stack((xvals, yvals)))
        scatter1.set_sizes(new_sizes)
        scatter1.set_facecolors(new_colors)
        scatter1.set_alpha(slider_alpha1.val)
        scatter2.set_offsets(np.column_stack((xvals, yvals)))
        scatter2.set_sizes(new_sizes)
        scatter2.set_facecolors(new_colors)
        scatter2.set_alpha(slider_alpha2.val)
        ax1.set_title(f"Indicator: {current_indicator}{title_suffix} (Left Plot)")
        ax2.set_title(f"Indicator: {current_indicator}{title_suffix} (Right Plot)")
        fig.canvas.draw_idle()

    def on_indicator_label_clicked(label):
        nonlocal current_indicator
        current_indicator = label
        update_scatter()

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 7))
    plt.subplots_adjust(left=0.25, bottom=0.35)
    updating_axes = [False]

    def on_xlim_changed(ax):
        if updating_axes[0]:
            return
        try:
            updating_axes[0] = True
            other_ax = ax2 if ax == ax1 else ax1
            other_ax.set_xlim(ax.get_xlim())
            fig.canvas.draw_idle()
        finally:
            updating_axes[0] = False

    def on_ylim_changed(ax):
        if updating_axes[0]:
            return
        try:
            updating_axes[0] = True
            other_ax = ax2 if ax == ax1 else ax1
            other_ax.set_ylim(ax.get_ylim())
            fig.canvas.draw_idle()
        finally:
            updating_axes[0] = False

    ax1.callbacks.connect('xlim_changed', on_xlim_changed)
    ax1.callbacks.connect('ylim_changed', on_ylim_changed)
    ax2.callbacks.connect('xlim_changed', on_xlim_changed)
    ax2.callbacks.connect('ylim_changed', on_ylim_changed)

    initial_data = get_current_data()
    if initial_data.get('geo_bounds'):
        geo_bounds = initial_data['geo_bounds']
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:3857", always_xy=True)
        x_min, y_min = transformer.transform(geo_bounds[0], geo_bounds[2])
        x_max, y_max = transformer.transform(geo_bounds[1], geo_bounds[3])
        image_extent = [x_min, x_max, y_min, y_max]
        ax1.set_xlim(image_extent[0], image_extent[1])
        ax1.set_ylim(image_extent[2], image_extent[3])
        ax2.set_xlim(image_extent[0], image_extent[1])
        ax2.set_ylim(image_extent[2], image_extent[3])
        ctx.add_basemap(ax1, crs="EPSG:3857", source=ctx.providers.OpenStreetMap.Mapnik)
        ctx.add_basemap(ax2, crs="EPSG:3857", source=ctx.providers.OpenStreetMap.Mapnik)
    else:
        image_extent = [0, 512, 0, 512]
        background_img = Image.open(initial_data['background'])
        ax1.imshow(background_img, extent=image_extent, aspect='auto')
        ax2.imshow(background_img, extent=image_extent, aspect='auto')

    scatter1 = ax1.scatter([], [], c='white')
    ax1.set_title(f"Indicator: {current_indicator} (Left Plot)")
    scatter2 = ax2.scatter([], [], c='white')
    ax2.set_title(f"Indicator: {current_indicator} (Right Plot)")

    ax_size = plt.axes([0.25, 0.2, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_min = plt.axes([0.25, 0.17, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_max = plt.axes([0.25, 0.14, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_alpha1 = plt.axes([0.25, 0.11, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_alpha2 = plt.axes([0.25, 0.08, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_exp = plt.axes([0.25, 0.05, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_sharp = plt.axes([0.25, 0.02, 0.65, 0.02], facecolor='lightgoldenrodyellow')

    slider_size = Slider(ax_size, 'Size', 1, 30, valinit=29.9)
    slider_min = Slider(ax_min, 'Min', 0.0, 1.0, valinit=0.838)
    slider_max = Slider(ax_max, 'Max', 1.0, 2.0, valinit=1.252)
    slider_alpha1 = Slider(ax_alpha1, 'Alpha1', 0.0, 1.0, valinit=0.687)
    slider_alpha2 = Slider(ax_alpha2, 'Alpha2', 0.0, 1.0, valinit=0.05)
    slider_exp = Slider(ax_exp, 'Exponent', -10, 0, valinit=-2.02)
    slider_sharp = Slider(ax_sharp, 'Sharp', 0.1, 2.0, valinit=1.192)

    ax_radio = plt.axes([0.05, 0.35, 0.15, 0.5], facecolor='lightgoldenrodyellow')
    radio_buttons = RadioButtons(ax_radio, indicators, active=0)
    radio_buttons.on_clicked(on_indicator_label_clicked)

    def slider_updated(val):
        update_scatter()
    slider_size.on_changed(slider_updated)
    slider_min.on_changed(slider_updated)
    slider_max.on_changed(slider_updated)
    slider_alpha1.on_changed(slider_updated)
    slider_alpha2.on_changed(slider_updated)
    slider_exp.on_changed(slider_updated)
    slider_sharp.on_changed(slider_updated)

    update_scatter()
    plt.show()

# ============================================================================
# Main Function: Process Indicators and Launch UI
# ============================================================================

def main_option2_single_window():
    """
    1) Scan for mother folders (e.g., "8", "pipe", etc.) in working_dir.
    2) For each, check for subfolders: <indicator>_grid, <indicator>_jpgs, <indicator>_result.
    3) If no grid file exists, prompt the user to create one interactively on the map.
    4) Process the grid and images if not already complete.
    5) If a KML file is present, use its coordinates to compute geo_points.
    6) Launch an interactive UI.
    """
    mother_folders = [folder for folder in os.listdir(working_dir) if os.path.isdir(os.path.join(working_dir, folder))]
    valid_folders = []
    for folder in mother_folders:
        candidate = os.path.join(working_dir, folder)
        grid_sub = os.path.join(candidate, f"{folder}_grid")
        jpgs_sub = os.path.join(candidate, f"{folder}_jpgs")
        result_sub = os.path.join(candidate, f"{folder}_result")
        if os.path.isdir(grid_sub) and os.path.isdir(jpgs_sub):
            os.makedirs(result_sub, exist_ok=True)
            valid_folders.append(folder)
    if not valid_folders:
        logging.error("No valid mother folders found with required subfolders.")
        return

    all_scatter_data = {}
    for indicator in valid_folders:
        mother_dir = os.path.join(working_dir, indicator)
        grid_dir = os.path.join(mother_dir, f"{indicator}_grid")
        if not find_file(grid_dir, r"grid_.*\.xlsx"):
            create = messagebox.askyesno("Create Grid", f"No grid file found for indicator '{indicator}'.\nDo you want to create one interactively on the map?")
            if create:
                grid_path = create_grid_on_map(indicator, os.path.join(mother_dir, f"{indicator}_jpgs"))
                if grid_path:
                    os.makedirs(grid_dir, exist_ok=True)
                    os.replace(grid_path, os.path.join(grid_dir, os.path.basename(grid_path)))
        jpgs_dir = os.path.join(mother_dir, f"{indicator}_jpgs")
        result_dir = os.path.join(mother_dir, f"{indicator}_result")
        if is_indicator_complete(result_dir, indicator):
            logging.info(f"Indicator '{indicator}' is already complete. Skipping re-run.")
        else:
            logging.info(f"Processing indicator: {indicator}")
            if not run_partA(grid_dir, jpgs_dir, result_dir, indicator):
                logging.error(f"Indicator {indicator}: Part A failed. Skipping.")
                continue
            run_partC(result_dir, indicator)
            run_partD(result_dir, indicator)
            run_partE(result_dir, indicator)
            chart_path = os.path.join(result_dir, f"chart_{indicator}.xlsx")
            force_recalc_with_excel(chart_path)
        grid_pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
        chart_pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
        grid_file = find_file(grid_dir, grid_pattern)
        chart_file = find_file(result_dir, chart_pattern)
        if not grid_file or not chart_file:
            logging.warning(f"Missing grid or chart file for indicator '{indicator}'.")
            continue
        points, _ = read_points_from_xlsx_openpyxl(grid_file)
        if not points:
            logging.warning(f"No points in grid file for indicator '{indicator}'. Skipping.")
            continue
        x_vals, y_vals = zip(*points)
        kml_file = find_file(mother_dir, r".*\.kml")
        kml_bounds = parse_kml_bounds(kml_file) if kml_file else None
        geo_points = None
        geo_bounds = None
        if kml_bounds:
            lon_min, lon_max, lat_min, lat_max = kml_bounds
            geo_bounds = [lon_min, lon_max, lat_min, lat_max]
            geo_x_vals = [lon_min + (x/512)*(lon_max - lon_min) for x in x_vals]
            geo_y_vals = [lat_min + (y/512)*(lat_max - lat_min) for y in y_vals]
            geo_points = (geo_x_vals, geo_y_vals)
        try:
            wb = load_workbook(chart_file, data_only=True)
            if 'Sheet3' not in wb.sheetnames:
                logging.error(f"Sheet3 missing in {chart_file}.")
                continue
            ws = wb['Sheet3']
            data_array = [float(row[0]) for row in ws.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True) if isinstance(row[0], (int, float))]
            wb.close()
        except Exception as e:
            logging.error(f"Error reading data from {chart_file}: {e}")
            continue
        frame0 = os.path.join(jpgs_dir, 'frame_0.jpg')
        if not os.path.isfile(frame0):
            cand = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
            if cand:
                frame0 = os.path.join(jpgs_dir, cand[0])
            else:
                logging.error(f"No background image in {jpgs_dir}. Skipping {indicator}.")
                continue
        all_scatter_data[indicator] = {
            'points': (list(x_vals), list(y_vals)),
            'geo_points': geo_points,
            'geo_bounds': geo_bounds,
            'data': np.array(data_array),
            'background': frame0
        }
    run_interactive_scatter_for_all_indicators(all_scatter_data)

if __name__ == '__main__':
    main_option2_single_window()
