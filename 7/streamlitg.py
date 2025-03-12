import streamlit as st
import os
import re
import logging
import numpy as np
import pandas as pd
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from multiprocessing import Pool, cpu_count
import matplotlib.pyplot as plt
import pytesseract
import json
import xml.etree.ElementTree as ET
from pyproj import Transformer
from matplotlib.path import Path as MplPath
from pathlib import Path

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Set Tesseract path (adjust for deployment environment)
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

# Reference dimensions
ref_width = 512
ref_height = 512

# Session state initialization
if 'step' not in st.session_state:
    st.session_state.step = 'upload'
if 'indicator' not in st.session_state:
    st.session_state.indicator = None
if 'grid_created' not in st.session_state:
    st.session_state.grid_created = False
if 'processing_complete' not in st.session_state:
    st.session_state.processing_complete = False
if 'all_data' not in st.session_state:
    st.session_state.all_data = {}
if 'available_indicators' not in st.session_state:
    st.session_state.available_indicators = []

# Helper Functions
def gif_to_frames(gif_path, output_folder):
    with Image.open(gif_path) as img:
        i = 0
        while True:
            try:
                img.seek(i)
                frame = img.convert('RGB')
                frame.save(os.path.join(output_folder, f'frame_{i}.jpg'), 'JPEG')
                i += 1
            except EOFError:
                break
    return True

def find_file(folder, pattern):
    regex = re.compile(pattern, re.IGNORECASE)
    for fname in os.listdir(folder):
        if regex.search(fname):
            return os.path.join(folder, fname)
    return None

def parse_kml_bounds(kml_path):
    try:
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        coords_text = root.find('.//kml:coordinates', ns).text.strip()
        coords = coords_text.split()
        unique_coords = []
        for coord in coords:
            parts = coord.split(',')
            if len(parts) >= 2:
                lon, lat = float(parts[0]), float(parts[1])
                if (lon, lat) not in unique_coords:
                    unique_coords.append((lon, lat))
        if len(unique_coords) < 4:
            return None
        return (unique_coords[0][0], unique_coords[1][0], unique_coords[3][1], unique_coords[0][1])
    except Exception as e:
        logging.error(f"Error parsing KML file {kml_path}: {e}")
        return None

def parse_txt_bounds(txt_path):
    try:
        with open(txt_path, 'r', encoding='utf-8-sig') as f:
            data = json.loads(f.read().strip())
        if data.get("type") == "Polygon" and "coordinates" in data:
            coords = data["coordinates"][0]
            lons = [pt[0] for pt in coords]
            lats = [pt[1] for pt in coords]
            return min(lons), max(lons), min(lats), max(lats)
        return None
    except Exception as e:
        logging.error(f"Error parsing txt file {txt_path}: {e}")
        return None

def read_rgb_values_at_point(image, point):
    try:
        pixel = image.getpixel((int(point[0]), int(point[1])))
        if isinstance(pixel, (int, float)):
            return (pixel, pixel, pixel)
        return pixel[:3]
    except Exception as e:
        logging.error(f"Error reading pixel at {point}: {e}")
        return None

def read_points_from_xlsx_openpyxl(file_path):
    points = []
    grid_shape = (None, None)
    wb = load_workbook(file_path, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    points = [(float(row[0]), float(row[1])) for row in sheet.iter_rows(min_row=2, values_only=True)
              if row[0] is not None and row[1] is not None]
    n_rows = sheet['D2'].value
    n_cols = sheet['E2'].value
    grid_shape = (int(n_rows) if n_rows else None, int(n_cols) if n_cols else None)
    wb.close()
    return points, grid_shape

def process_image(image_path, points):
    img = Image.open(image_path).convert("RGB")
    img_width, img_height = img.size
    scaled_points = [(xx * img_width / ref_width, yy * img_height / ref_height) for xx, yy in points]
    results = []
    for pt in scaled_points:
        rgb = read_rgb_values_at_point(img, pt)
        if rgb:
            r, g, b = rgb
            grayscale = 0.299 * r + 0.587 * g + 0.114 * b
            results.append([pt[0], pt[1], round(r, 3), round(g, 3), round(b, 3), round(grayscale, 3)])
    img.close()
    return results

def write_results_to_excel(results, workbook_path):
    wb = Workbook()
    wb.remove(wb.active)
    for index, data in enumerate(results):
        sheet = wb.create_sheet(title=str(index))
        sheet.append(["X", "Y", "R", "G", "B", "Grayscale"])
        for row in data:
            sheet.append(row)
    wb.save(workbook_path)

def extract_date(image_path):
    config = "--psm 7 -c tessedit_char_whitelist=0123456789-"
    img = Image.open(image_path)
    width, height = img.size
    roi_dyn = img.crop((int(0.75 * width), 0, width, int(0.15 * height)))
    roi_dyn = roi_dyn.resize((roi_dyn.width * 2, roi_dyn.height * 2), Image.LANCZOS)
    dyn_images = pil_multi_threshold_preprocess(roi_dyn)
    date_dyn = pil_attempt_ocr(dyn_images, config)
    if date_dyn:
        return date_dyn
    roi_fixed = img.crop((width - 160, 0, width, 60))
    roi_fixed = roi_fixed.resize((roi_fixed.width * 2, roi_fixed.height * 2), Image.LANCZOS)
    fixed_images = pil_multi_threshold_preprocess(roi_fixed)
    return pil_attempt_ocr(fixed_images, config)

def pil_multi_threshold_preprocess(roi, thresholds=[80, 120, 160, 200], invert=True):
    gray = roi.convert("L")
    for thr in thresholds:
        bin_img = gray.point(lambda p: 255 if p > thr else 0)
        yield bin_img
        if invert:
            yield ImageOps.invert(bin_img)

def pil_attempt_ocr(preprocessed_images, config):
    for img in preprocessed_images:
        text = pytesseract.image_to_string(img, config=config).strip()
        found = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        if found:
            try:
                return datetime.strptime(found[0], "%Y-%m-%d")
            except:
                continue
    return None

def save_ocr_results(jpgs_dir):
    folder_path = Path(jpgs_dir)
    data = []
    for filename in sorted(folder_path.iterdir(), key=lambda x: extract_number(x)):
        if filename.suffix.lower() in ('.jpg', '.jpeg'):
            frame_num = extract_number(filename)
            dt = extract_date(filename)
            data.append((frame_num, dt))
    wb = Workbook()
    ws = wb.active
    ws.title = "OCR_Dates"
    ws.append(["FrameNumber", "RecognizedDate"])
    for fnum, dt in data:
        dt_str = dt.strftime("%Y-%m-%d") if dt else ""
        ws.append([fnum, dt_str])
    out_path = folder_path / "ocr_results.xlsx"
    wb.save(out_path)
    return out_path

def extract_number(filename):
    number = re.findall(r'\d+$', filename.stem)
    return int(number[0]) if number else 0

def is_indicator_complete(grid_dir, result_dir, indicator):
    grid_file = find_file(grid_dir, rf"grid_.*{re.escape(indicator)}.*\.xlsx")
    results_file = find_file(result_dir, rf"results_{re.escape(indicator)}\.xlsx")
    chart_file = find_file(result_dir, rf"chart_.*{re.escape(indicator)}.*\.xlsx")
    jpgs_dir = grid_dir.replace("_grid", "_jpgs")
    jpgs_exist = any(f.lower().endswith('.jpg') for f in os.listdir(jpgs_dir)) if os.path.exists(jpgs_dir) else False
    return all([grid_file, results_file, chart_file, jpgs_exist]) and \
           os.path.exists(grid_file) and os.path.exists(results_file) and os.path.exists(chart_file)

def scan_existing_indicators(base_dir):
    if not os.path.exists(base_dir):
        return []
    indicators = []
    for folder in os.listdir(base_dir):
        if folder.endswith('_grid'):
            indicator = folder.replace('_grid', '')
            grid_dir = os.path.join(base_dir, folder)
            result_dir = os.path.join(base_dir, f"{indicator}_result")
            if os.path.exists(result_dir) and is_indicator_complete(grid_dir, result_dir, indicator):
                indicators.append(indicator)
    return indicators

def rotate_points(points, angle_deg, center):
    angle_rad = np.radians(angle_deg)
    cos_a, sin_a = np.cos(angle_rad), np.sin(angle_rad)
    return [[(pt[0] - center[0]) * cos_a - (pt[1] - center[1]) * sin_a + center[0],
             (pt[0] - center[0]) * sin_a + (pt[1] - center[1]) * cos_a + center[1]] for pt in points]

def preview_grid(bg_img, grid_pts, poly_arr=None):
    fig, ax = plt.subplots()
    ax.imshow(bg_img, extent=[0, ref_width, ref_height, 0], origin='upper')
    if grid_pts:
        xs, ys = zip(*grid_pts)
        ax.scatter(xs, ys, c='red', s=10, label='Grid Points')
    if poly_arr:
        poly_arr = np.append(poly_arr, [poly_arr[0]], axis=0)
        ax.plot(poly_arr[:, 0], poly_arr[:, 1], color='blue', linestyle='--', label='Polygon')
    ax.legend()
    st.pyplot(fig)

def save_grid(grid_pts, grid_dir, indicator, rows=None, cols=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "GridPoints"
    ws.append(["X", "Y"])
    for pt in grid_pts:
        ws.append(pt)
    if rows and cols:
        ws['D1'] = "Rows"
        ws['D2'] = rows
        ws['E1'] = "Cols"
        ws['E2'] = cols
    else:
        ys = [pt[1] for pt in grid_pts]
        unique_y = np.unique(np.round(ys, 2))
        n_rows = len(unique_y)
        n_cols = int(np.ceil(len(grid_pts) / n_rows))
        ws['D1'] = "Rows"
        ws['D2'] = n_rows
        ws['E1'] = "Cols"
        ws['E2'] = n_cols
    grid_file = os.path.join(grid_dir, f"grid_{indicator}.xlsx")
    wb.save(grid_file)

def run_partC(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.endswith(".xlsx") and not filename.startswith("~$"):
            input_file = os.path.join(output_folder, filename)
            match = re.search(r'_([A-Za-z0-9]+)\.xlsx$', filename)
            if not match or match.group(1) != indicator:
                continue
            wb = load_workbook(input_file, data_only=True)
            data_frames = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = list(sheet.values)[1:]
                df = pd.DataFrame(data)
                col_data = df.iloc[:, 3].dropna()
                if not col_data.empty:
                    data_frames.append(col_data.reset_index(drop=True).to_frame(name=sheet_name))
            if data_frames:
                all_data_df = pd.concat(data_frames, axis=1)
                all_data_df['Average'] = all_data_df.mean(axis=1)
                out_file = os.path.join(output_folder, f"total_results_{indicator}.xlsx")
                with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                    all_data_df.to_excel(writer, sheet_name='ChlA', index=False)
                wb.close()

def run_partD(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("total_results_") and filename.endswith(".xlsx"):
            source_file = os.path.join(output_folder, filename)
            wb = load_workbook(source_file, data_only=True)
            ws = wb['ChlA']
            avg_col = next((col for col in range(1, ws.max_column + 1) if ws.cell(row=1, column=col).value == 'Average'), None)
            if avg_col:
                avg_data = [ws.cell(row=r, column=avg_col).value for r in range(2, ws.max_row + 1)]
                grid_file = find_file(st.session_state.grid_dir, rf"grid_.*{re.escape(indicator)}.*\.xlsx")
                _, grid_shape = read_points_from_xlsx_openpyxl(grid_file)
                grid_rows = grid_shape[0] or 20
                wb_out = Workbook()
                ws1 = wb_out.active
                ws1.title = 'Sheet1'
                for r in range(1, grid_rows + 1):
                    ws1.cell(row=r, column=2, value=r).number_format = "0.0000"
                col_index = 3
                for i in range(0, len(avg_data), grid_rows):
                    for j in range(grid_rows):
                        if i + j < len(avg_data):
                            ws1.cell(row=j + 1, column=col_index, value=avg_data[i + j]).number_format = "0.0000"
                    col_index += 1
                ws2 = wb_out.create_sheet(title='Sheet2')
                for col in range(3, col_index):
                    col_letter = get_column_letter(col)
                    for row in range(1, grid_rows + 1):
                        ws2.cell(row=row, column=col, value=f"=Sheet1!{col_letter}{row}/AVERAGE(Sheet1!{col_letter}$1:{col_letter}${grid_rows})")
                final_chart = os.path.join(output_folder, f"chart_{indicator}.xlsx")
                wb_out.save(final_chart)

def load_existing_data(indicator, grid_dir, jpgs_dir, result_dir):
    grid_file = find_file(grid_dir, rf"grid_.*{re.escape(indicator)}.*\.xlsx")
    chart_file = find_file(result_dir, rf"chart_.*{re.escape(indicator)}.*\.xlsx")
    points, _ = read_points_from_xlsx_openpyxl(grid_file)
    x_vals, y_vals = zip(*points)
    wb = load_workbook(chart_file, data_only=True)
    ws = wb['Sheet3']
    data_array = [float(row[0]) for row in ws.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True) if isinstance(row[0], (int, float))]
    frame0 = find_file(jpgs_dir, r"frame_0\.jpg") or next((os.path.join(jpgs_dir, f) for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')), None)
    results_file = find_file(result_dir, f"results_{indicator}.xlsx")

    st.session_state.all_data[indicator] = {
        'points': (list(x_vals), list(y_vals)),
        'data': np.array(data_array),
        'background': frame0,
        'results_file': results_file
    }

# Streamlit Main Application
def main():
    st.title("Image Processing and Grid Analysis Application")
    st.write("""
        This application processes images based on a user-defined grid, performs OCR to extract dates,
        and generates analytical outputs. If existing data is detected, you may select a case to visualize directly.
        Otherwise, proceed with uploading new files and creating a grid.
    """)

    base_dir = os.path.join(os.getcwd(), "temp_processing")
    st.session_state.available_indicators = scan_existing_indicators(base_dir)

    if st.session_state.step == 'upload':
        st.header("Step 1: Upload Input Files or Select Existing Case")
        st.write("Select an existing case to visualize, or upload new files to begin processing.")

        if st.session_state.available_indicators:
            st.subheader("Existing Cases Detected")
            selected_indicator = st.selectbox(
                "Select an existing indicator to proceed with visualization:",
                options=st.session_state.available_indicators + ["New Case"],
                index=len(st.session_state.available_indicators)  # Default to "New Case"
            )
            if selected_indicator != "New Case" and st.button("Proceed with Selected Case"):
                st.session_state.indicator = selected_indicator
                st.session_state.grid_dir = os.path.join(base_dir, f"{selected_indicator}_grid")
                st.session_state.jpgs_dir = os.path.join(base_dir, f"{selected_indicator}_jpgs")
                st.session_state.result_dir = os.path.join(base_dir, f"{selected_indicator}_result")
                st.session_state.step = 'visualize'
                load_existing_data(selected_indicator, st.session_state.grid_dir, st.session_state.jpgs_dir, st.session_state.result_dir)
                st.success(f"Proceeding to visualization for indicator '{selected_indicator}'.")
                return

        st.subheader("Upload New Files")
        indicator = st.text_input("Enter Indicator Name for New Case (e.g., ChlA):", value="Indicator")
        uploaded_gif = st.file_uploader("Upload GIF (optional)", type=["gif"])
        uploaded_images = st.file_uploader("Upload JPG Images (optional)", type=["jpg", "jpeg"], accept_multiple_files=True)
        uploaded_bounds = st.file_uploader("Upload Bounds File (KML or TXT, optional)", type=["kml", "txt"])

        if st.button("Proceed with New Case"):
            st.session_state.indicator = indicator
            grid_dir = os.path.join(base_dir, f"{indicator}_grid")
            jpgs_dir = os.path.join(base_dir, f"{indicator}_jpgs")
            result_dir = os.path.join(base_dir, f"{indicator}_result")
            os.makedirs(base_dir, exist_ok=True)
            os.makedirs(grid_dir, exist_ok=True)
            os.makedirs(jpgs_dir, exist_ok=True)
            os.makedirs(result_dir, exist_ok=True)

            # Handle uploads
            uploaded = False
            if uploaded_gif:
                gif_path = os.path.join(jpgs_dir, "input.gif")
                with open(gif_path, "wb") as f:
                    f.write(uploaded_gif.read())
                gif_to_frames(gif_path, jpgs_dir)
                uploaded = True
            if uploaded_images:
                for img in uploaded_images:
                    with open(os.path.join(jpgs_dir, img.name), "wb") as f:
                        f.write(img.read())
                uploaded = True
            if uploaded_bounds:
                bounds_path = os.path.join(base_dir, uploaded_bounds.name)
                with open(bounds_path, "wb") as f:
                    f.write(uploaded_bounds.read())

            st.session_state.grid_dir = grid_dir
            st.session_state.jpgs_dir = jpgs_dir
            st.session_state.result_dir = result_dir

            if not uploaded and not any(f.lower().endswith('.jpg') for f in os.listdir(jpgs_dir)):
                st.error("No JPG images found or uploaded. Please upload images to proceed.")
                return

            # Check if all files exist
            if is_indicator_complete(grid_dir, result_dir, indicator) and not uploaded:
                st.success(f"All required files for indicator '{indicator}' already exist. Proceeding to visualization.")
                st.session_state.step = 'visualize'
                load_existing_data(indicator, grid_dir, jpgs_dir, result_dir)
            elif uploaded or not find_file(grid_dir, rf"grid_.*{re.escape(indicator)}.*\.xlsx"):
                st.session_state.step = 'grid'
                st.success("Files uploaded or grid missing. Proceeding to grid creation.")
            else:
                st.session_state.step = 'process'
                st.success("Some files exist, but processing is required. Proceeding to processing.")

    elif st.session_state.step == 'grid':
        st.header("Step 2: Define the Grid")
        st.write("""
            Define the grid by either drawing a polygon or setting an origin point with parameters.
            Adjust the sliders to customize the grid layout.
        """)

        jpgs_dir = st.session_state.jpgs_dir
        bg_path = find_file(jpgs_dir, r"frame_0\.jpg") or next((os.path.join(jpgs_dir, f) for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')), None)

        if not bg_path or not os.path.exists(bg_path):
            st.error("No background image found. Please ensure JPGs are available.")
            return

        bg_img = Image.open(bg_path)
        st.image(bg_img, caption="Background Image", use_column_width=True)

        mode = st.radio("Grid Creation Mode", ["Polygon", "Origin"])

        if mode == "Polygon":
            st.write("Define the polygon by specifying vertices in the format: x1,y1;x2,y2;...")
            poly_input = st.text_input("Polygon Vertices (e.g., 100,100;200,200;300,100):")
            h_spacing = st.slider("Horizontal Spacing", 1, 100, 50)
            v_spacing = st.slider("Vertical Spacing", 1, 100, 50)
            poly_rot = st.slider("Polygon Rotation (°)", -180, 180, 0)
            grid_rot = st.slider("Grid Rotation (°)", -180, 180, 0)

            if poly_input:
                try:
                    points = [tuple(map(float, pt.split(','))) for pt in poly_input.split(';')]
                    if len(points) < 3:
                        st.error("Polygon must have at least 3 points.")
                        return
                    centroid = np.mean(points, axis=0)
                    poly_arr = rotate_points(points, poly_rot, centroid)
                    poly_path = MplPath(poly_arr)
                    xs, ys = zip(*poly_arr)
                    bx_min, bx_max = min(xs), max(xs)
                    by_min, by_max = min(ys), max(ys)
                    pts = [(xx, yy) for yy in np.arange(by_min, by_max + v_spacing, v_spacing)
                           for xx in np.arange(bx_min, bx_max + h_spacing, h_spacing)]
                    grid_center = [(bx_min + bx_max) / 2, (by_min + by_max) / 2]
                    pts_rot = rotate_points(pts, grid_rot, grid_center)
                    grid_pts = [pt for pt in pts_rot if poly_path.contains_point(pt)]
                    grid_pts.sort(key=lambda pt: (pt[1], pt[0]))
                    preview_grid(bg_img, grid_pts, poly_arr)
                except Exception as e:
                    st.error(f"Error processing polygon: {e}")
                    return

        else:  # Origin Mode
            origin_x = st.slider("Origin X", 0, ref_width, 100)
            origin_y = st.slider("Origin Y", 0, ref_height, 100)
            spacing = st.slider("Spacing", 10, 200, 50)
            rows = st.slider("Rows", 1, 50, 10)
            cols = st.slider("Columns", 1, 50, 10)
            angle = st.slider("Angle (°)", -90, 90, 0)

            angle_rad = np.radians(angle)
            grid_pts = []
            for i in range(rows):
                for j in range(cols):
                    xx = j * spacing
                    yy = i * spacing
                    x_rot = xx * np.cos(angle_rad) - yy * np.sin(angle_rad)
                    y_rot = xx * np.sin(angle_rad) + yy * np.cos(angle_rad)
                    grid_pts.append([origin_x + x_rot, origin_y + y_rot])
            grid_pts.sort(key=lambda pt: (pt[1], pt[0]))
            preview_grid(bg_img, grid_pts)

        if st.button("Save Grid"):
            save_grid(grid_pts, st.session_state.grid_dir, st.session_state.indicator, rows if mode == "Origin" else None, cols if mode == "Origin" else None)
            st.session_state.grid_created = True
            st.session_state.step = 'process'
            st.success("Grid saved successfully. Proceeding to processing.")

    elif st.session_state.step == 'process':
        st.header("Step 3: Process Images")
        st.write("Processing images based on the defined grid. This may take some time if required.")
        
        grid_dir = st.session_state.grid_dir
        jpgs_dir = st.session_state.jpgs_dir
        result_dir = st.session_state.result_dir
        indicator = st.session_state.indicator

        if is_indicator_complete(grid_dir, result_dir, indicator):
            st.info(f"All required files for indicator '{indicator}' already exist. Proceeding to visualization.")
            st.session_state.step = 'visualize'
            load_existing_data(indicator, grid_dir, jpgs_dir, result_dir)
        else:
            grid_file = find_file(grid_dir, rf"grid_.*{re.escape(indicator)}.*\.xlsx")
            points, _ = read_points_from_xlsx_openpyxl(grid_file)
            jpg_files = [os.path.join(jpgs_dir, f) for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]

            if st.button("Start Processing"):
                with st.spinner("Processing..."):
                    with Pool(cpu_count()) as pool:
                        results = pool.starmap(process_image, [(jpg, points) for jpg in jpg_files])
                    output_xlsx = os.path.join(result_dir, f"results_{indicator}.xlsx")
                    write_results_to_excel(results, output_xlsx)
                    run_partC(result_dir, indicator)
                    run_partD(result_dir, indicator)
                    st.session_state.processing_complete = True
                    st.session_state.step = 'visualize'
                    st.success(f"Processing completed. Results saved to {output_xlsx}")
                    load_existing_data(indicator, grid_dir, jpgs_dir, result_dir)

    elif st.session_state.step == 'visualize':
        st.header("Step 4: Visualize Results")
        st.write("Interact with the processed data using the controls below.")

        indicator = st.session_state.indicator
        if indicator not in st.session_state.all_data:
            st.error("Data not loaded properly. Please restart the application.")
            return

        d = st.session_state.all_data[indicator]
        bg_img = Image.open(d['background'])
        w, h = bg_img.size
        xvals_scaled = np.array(d['points'][0]) * (w / ref_width)
        yvals_scaled = np.array(d['points'][1]) * (h / ref_height)

        size = st.slider("Point Size", 1, 30, 10)
        min_val = st.slider("Min Value", 0.0, 1.0, 0.5)
        max_val = st.slider("Max Value", 1.0, 2.0, 1.5)
        alpha = st.slider("Opacity", 0.0, 1.0, 0.5)

        fig, ax = plt.subplots()
        ax.imshow(bg_img, extent=[0, w, h, 0], origin='upper')
        scatter = ax.scatter(xvals_scaled, yvals_scaled, s=size, c=d['data'], cmap='gray', alpha=alpha, vmin=min_val, vmax=max_val)
        plt.colorbar(scatter, label='Grayscale Value')
        st.pyplot(fig)

if __name__ == "__main__":
    main()