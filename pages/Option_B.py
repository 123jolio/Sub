import streamlit as st
import base64, io
# Patch: Import streamlit.elements.image and add image_to_url accepting extra arguments.
import streamlit.elements.image as st_image
def image_to_url(img, *args, **kwargs):
    buffer = io.BytesIO()
    img.save(buffer, format="PNG")
    encoded_image = base64.b64encode(buffer.getvalue()).decode()
    return "data:image/png;base64," + encoded_image
st_image.image_to_url = image_to_url

import sys, os, re, logging, time, numpy as np, pandas as pd, json, math
from datetime import datetime
from multiprocessing import Pool, cpu_count, freeze_support
from pathlib import Path
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import xml.etree.ElementTree as ET
import contextily as ctx
from pyproj import Transformer
from matplotlib.path import Path as MplPath
import pytesseract
import plotly.graph_objects as go

# Import drawable canvas component (install via pip install streamlit-drawable-canvas)
from streamlit_drawable_canvas import st_canvas

# Set your Tesseract path if needed.
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

logging.basicConfig(level=logging.INFO)

# Define reference dimensions for the grid (originally 512×512)
ref_width = 512
ref_height = 512

#############################################
# GIF CONVERSION HELPER
#############################################
def gif_to_frames(gif_path, output_folder):
    """Convert a GIF into individual JPEG frames saved in output_folder."""
    with Image.open(gif_path) as img:
        i = 0
        while True:
            try:
                img.seek(i)
                frame = img.convert('RGB')
                frame.save(os.path.join(output_folder, f'frame_{i}.jpg'), 'JPEG')
                i += 1
            except EOFError:
                break
    logging.info("GIF has been processed. Frames saved as JPGs in: " + output_folder)

#############################################
# HELPER FUNCTIONS
#############################################
def find_file(folder, pattern):
    """Return the first file in folder matching the regex pattern (case-insensitive)."""
    regex = re.compile(pattern, re.IGNORECASE)
    for fname in os.listdir(folder):
        if regex.search(fname):
            return os.path.join(folder, fname)
    return None

def parse_kml_bounds(kml_path):
    """Extract (lon_min, lon_max, lat_min, lat_max) from a KML file."""
    try:
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        coords_text = root.find('.//kml:coordinates', ns).text.strip()
        coords = coords_text.split()
        unique_coords = []
        for coord in coords:
            parts = coord.split(',')
            if len(parts) >= 2:
                lon, lat = float(parts[0]), float(parts[1])
                if (lon, lat) not in unique_coords:
                    unique_coords.append((lon, lat))
        if len(unique_coords) < 4:
            return None
        return (unique_coords[0][0], unique_coords[1][0],
                unique_coords[3][1], unique_coords[0][1])
    except Exception as e:
        logging.error(f"Error parsing KML file {kml_path}: {e}")
        return None

def parse_txt_bounds(txt_path):
    """
    Extract (lon_min, lon_max, lat_min, lat_max) from a TXT file containing JSON data.
    Expected format: {"type":"Polygon","coordinates":[[[lon,lat],[lon,lat], ...]]}
    """
    try:
        with open(txt_path, 'r', encoding='utf-8-sig') as f:
            content = f.read().strip()
            if not content:
                logging.error(f"File {txt_path} is empty!")
                return None
            data = json.loads(content)
        if data.get("type") == "Polygon" and "coordinates" in data:
            coords = data["coordinates"][0]
            lons = [pt[0] for pt in coords]
            lats = [pt[1] for pt in coords]
            return min(lons), max(lons), min(lats), max(lats)
        else:
            logging.error(f"JSON in {txt_path} does not have expected keys.")
            return None
    except Exception as e:
        logging.error(f"Error parsing txt file {txt_path}: {e}")
        return None

def set_calculation_properties(workbook):
    try:
        cp = workbook.calculation_properties
        cp.calculationMode = "auto"
        cp.calcCompleted = False
        cp.calcOnSave = True
        cp.fullCalcOnLoad = True
        cp.forceFullCalc = True
    except AttributeError:
        logging.warning("Calculation properties not supported in this openpyxl version.")

def force_recalc_with_excel(final_file):
    if not os.path.isfile(final_file):
        logging.error(f"Cannot recalc: File not found: {final_file}")
        return
    try:
        import win32com.client as win32
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(final_file, UpdateLinks=False)
        wb.Application.CalculateFullRebuild()
        while wb.Application.CalculationState != 0:
            time.sleep(0.5)
        wb.Save()
        wb.Close(False)
        excel.Quit()
        logging.info(f"Excel recalculation forced and file saved: {final_file}")
    except Exception as e:
        logging.error(f"Error during Excel recalculation: {e}")

def is_indicator_complete(result_folder, indicator):
    pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
    return bool(find_file(result_folder, pattern))

#############################################
# READ RGB VALUES HELPER FUNCTION
#############################################
def read_rgb_values_at_point(image, point):
    try:
        pixel = image.getpixel((int(point[0]), int(point[1])))
        if isinstance(pixel, (int, float)):
            return (pixel, pixel, pixel)
        return pixel[:3]
    except Exception as e:
        logging.error(f"Error reading pixel at {point}: {e}")
        return None

#############################################
# READING GRID AND RESULTS
#############################################
def read_points_from_xlsx_openpyxl(file_path):
    points = []
    grid_shape = (None, None)
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        points = [(float(row[0]), float(row[1])) for row in sheet.iter_rows(min_row=2, values_only=True)
                  if row[0] is not None and row[1] is not None]
        n_rows = sheet['D2'].value
        n_cols = sheet['E2'].value
        grid_shape = (int(n_rows) if n_rows is not None else None,
                      int(n_cols) if n_cols is not None else None)
    except Exception as e:
        logging.error(f"Error reading grid file {file_path}: {e}")
    finally:
        wb.close()
    return points, grid_shape

def get_time_series_for_points(results_file, indices):
    wb = load_workbook(results_file, data_only=True)
    time_series = {idx: [] for idx in indices}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx in indices:
            try:
                val = float(rows[idx][5])
            except (IndexError, TypeError):
                val = None
            time_series[idx].append(val)
    wb.close()
    return time_series

#############################################
# OCR FUNCTIONS
#############################################
def pil_multi_threshold_preprocess(roi, thresholds=[80,120,160,200], invert=True):
    gray = roi.convert("L")
    for thr in thresholds:
        bin_img = gray.point(lambda p: 255 if p > thr else 0)
        yield bin_img
        if invert:
            yield ImageOps.invert(bin_img)

def pil_attempt_ocr(preprocessed_images, config):
    for img in preprocessed_images:
        text = pytesseract.image_to_string(img, config=config).strip()
        found = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        if found:
            dt_str = found[0]
            try:
                return datetime.strptime(dt_str, "%Y-%m-%d")
            except:
                continue
    return None

def extract_date(image_path):
    config = "--psm 7 -c tessedit_char_whitelist=0123456789-"
    try:
        img = Image.open(str(image_path))
    except Exception as e:
        st.error(f"Error: Unable to load image at {image_path}: {e}")
        return None
    width, height = img.size
    roi_dyn = img.crop((int(0.75*width), 0, width, int(0.15*height)))
    roi_dyn = roi_dyn.resize((roi_dyn.width*2, roi_dyn.height*2), Image.LANCZOS)
    dyn_images = pil_multi_threshold_preprocess(roi_dyn, thresholds=[80,120,160,200], invert=True)
    date_dyn = pil_attempt_ocr(dyn_images, config)
    if date_dyn:
        st.write(f"Dynamic OCR from {Path(image_path).name}: recognized {date_dyn.strftime('%Y-%m-%d')}")
        return date_dyn
    else:
        st.write(f"Dynamic OCR failed for {Path(image_path).name}, attempting fixed cropping.")
    roi_fixed = img.crop((width-160, 0, width, 60))
    roi_fixed = roi_fixed.resize((roi_fixed.width*2, roi_fixed.height*2), Image.LANCZOS)
    fixed_images = pil_multi_threshold_preprocess(roi_fixed, thresholds=[80,120,160,200], invert=True)
    date_fixed = pil_attempt_ocr(fixed_images, config)
    if date_fixed:
        st.write(f"Fixed OCR from {Path(image_path).name}: recognized {date_fixed.strftime('%Y-%m-%d')}")
        return date_fixed
    else:
        st.write(f"Date not found in {Path(image_path).name}")
        return None

def extract_number(filename):
    number = re.findall(r'\d+$', filename.stem)
    return int(number[0]) if number else 0

#############################################
# OCR RESULTS SAVING & READING
#############################################
def save_ocr_results(jpgs_dir):
    folder_path = Path(jpgs_dir)
    data = []
    for filename in sorted(folder_path.iterdir()):
        if filename.suffix.lower() in ('.png','.jpg','.jpeg'):
            frame_num = extract_number(filename)
            dt = extract_date(filename)
            data.append((frame_num, dt))
    data.sort(key=lambda x: x[0])
    wb_ocr = Workbook()
    ws = wb_ocr.active
    ws.title = "OCR_Dates"
    ws.append(["FrameNumber", "RecognizedDate"])
    for (fnum, dt) in data:
        dt_str = dt.strftime("%Y-%m-%d") if dt else ""
        ws.append([fnum, dt_str])
    out_path = folder_path / "ocr_results.xlsx"
    wb_ocr.save(str(out_path))
    st.write(f"OCR results saved to {out_path}")
    return out_path

def read_ocr_dates(jpgs_dir):
    folder_path = Path(jpgs_dir)
    ocr_file = folder_path / "ocr_results.xlsx"
    if not ocr_file.exists():
        ocr_file = save_ocr_results(jpgs_dir)
    wb = load_workbook(str(ocr_file), data_only=True)
    ws = wb.active
    frame_dates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        frame_num = row[0]
        date_str = row[1]
        dt = None
        if date_str:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
            except:
                pass
        frame_dates[frame_num] = dt
    wb.close()
    if frame_dates:
        max_frame = max(frame_dates.keys())
        out = [None]*(max_frame+1)
        for i in range(max_frame+1):
            out[i] = frame_dates.get(i, None)
        return out
    else:
        return []

#############################################
# GRID CREATION FUNCTIONS (Dynamic with Drawable Canvas)
#############################################
def streamlit_create_grid_on_map(indicator, jpgs_dir):
    st.header(f"Interactive Grid Creation for Indicator: {indicator}")
    # Display background image
    jpg_candidates = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
    if not jpg_candidates:
         st.error("No background image found in " + jpgs_dir)
         return None
    bg_path = os.path.join(jpgs_dir, "frame_0.jpg") if os.path.isfile(os.path.join(jpgs_dir, "frame_0.jpg")) else os.path.join(jpgs_dir, jpg_candidates[0])
    bg_img = Image.open(bg_path)
    st.image(bg_img, caption="Background Image", use_column_width=True)
    
    st.write("**Click on the image to add grid points.**")
    canvas_result = st_canvas(
        fill_color="rgba(0, 0, 0, 0)",  # transparent fill
        stroke_width=5,
        stroke_color="red",
        background_image=bg_img,
        update_streamlit=True,
        height=bg_img.height,
        width=bg_img.width,
        drawing_mode="point",
        key="canvas_grid",
    )
    
    points = []
    if canvas_result.json_data is not None:
        objects = canvas_result.json_data.get("objects", [])
        for obj in objects:
            if obj["type"] == "circle":
                x = obj["left"]
                y = obj["top"]
                points.append([x, y])
        st.write("Selected grid points:", points)
    
    if not points:
        st.warning("No points selected yet. Please click on the image to set grid points.")
        return None
    
    # Save grid points to Excel
    points.sort(key=lambda pt: (pt[1], pt[0]))
    wb = Workbook()
    ws = wb.active
    ws.title = "GridPoints"
    ws.append(["X", "Y"])
    for pt in points:
         ws.append(pt)
    unique_y = np.unique(np.round([pt[1] for pt in points], 2))
    n_rows = len(unique_y)
    n_cols = int(np.ceil(len(points) / n_rows)) if n_rows > 0 else 0
    ws['D1'] = "Rows"
    ws['D2'] = n_rows
    ws['E1'] = "Cols"
    ws['E2'] = n_cols
    grid_filename = f"grid_{indicator}.xlsx"
    save_path = os.path.join(os.getcwd(), grid_filename)
    wb.save(save_path)
    st.success(f"Grid file created: {save_path}")
    return save_path

#############################################
# PART A: Process Images Based on Grid
#############################################
def process_image(image_path, points):
    img = Image.open(image_path).convert("RGB")
    img_width, img_height = img.size
    scaled_points = [
        (xx * img_width / ref_width,
         yy * img_height / ref_height)
        for xx, yy in points
    ]
    results = []
    for pt in scaled_points:
        rgb = read_rgb_values_at_point(img, pt)
        if rgb:
            r, g, b = rgb
            grayscale = 0.299*r + 0.587*g + 0.114*b
            results.append([pt[0], pt[1], round(r, 3), round(g, 3), round(b, 3), round(grayscale, 3)])
    img.close()
    return results

def write_results_to_excel(results, workbook_path):
    wb = Workbook()
    wb.remove(wb.active)
    for index, data in enumerate(results):
        sheet = wb.create_sheet(title=str(index))
        sheet.append(["X", "Y", "R", "G", "B", "Grayscale"])
        for row in data:
            sheet.append(row)
    set_calculation_properties(wb)
    wb.save(workbook_path)

def run_partA(input_grid_folder, input_images_folder, output_folder, indicator):
    pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
    grid_file = find_file(input_grid_folder, pattern)
    if not grid_file:
        logging.error(f"Grid file not found in {input_grid_folder} for indicator {indicator}")
        return False
    points, _ = read_points_from_xlsx_openpyxl(grid_file)
    if not points:
        logging.error(f"No points read from {grid_file}. Exiting Part A.")
        return False
    jpg_files = [os.path.join(input_images_folder, f)
                 for f in os.listdir(input_images_folder)
                 if f.lower().endswith('.jpg')]
    if not jpg_files:
        logging.error(f"No .jpg files found in {input_images_folder}.")
        return False
    with Pool(cpu_count()) as pool:
        results = pool.starmap(process_image, [(jpg, points) for jpg in jpg_files])
    output_xlsx = os.path.join(output_folder, f"results_{indicator}.xlsx")
    write_results_to_excel(results, output_xlsx)
    logging.info(f"Part A completed. Wrote {output_xlsx}")
    return True

#############################################
# PART C
#############################################
def run_partC(output_folder, indicator):
    import pandas as pd
    for filename in os.listdir(output_folder):
        if filename.startswith("~$") or filename.startswith("chart_"):
            continue
        if filename.endswith(".xlsx"):
            input_file = os.path.join(output_folder, filename)
            match = re.search(r'_([A-Za-z0-9]+)\.xlsx$', filename)
            if not match:
                continue
            file_indicator = match.group(1)
            out_file = os.path.join(output_folder, f"total_results_{file_indicator}.xlsx")
            wb = load_workbook(input_file, data_only=True)
            data_frames = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = list(sheet.values)
                if not data:
                    continue
                data = data[1:]
                df = pd.DataFrame(data)
                try:
                    col_data = df.iloc[:, 3].dropna()
                    if not col_data.empty:
                        data_frames.append(col_data.reset_index(drop=True).to_frame(name=sheet_name))
                except IndexError:
                    logging.info(f"No Column 'D' in {sheet_name}")
            if data_frames:
                all_data_df = pd.concat(data_frames, axis=1)
                all_data_df['Average'] = all_data_df.mean(axis=1)
                with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                    all_data_df.to_excel(writer, sheet_name='ChlA', index=False)
                    set_calculation_properties(writer.book)
                logging.info(f"Part C: Wrote {out_file}")

#############################################
# PART D
#############################################
def run_partD(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("~$"):
            continue
        if filename.startswith("total_results_") and filename.endswith(".xlsx"):
            source_file = os.path.join(output_folder, filename)
            file_indicator = indicator
            wb = load_workbook(source_file, data_only=True)
            if 'ChlA' not in wb.sheetnames:
                logging.info(f"Sheet 'ChlA' not found in {filename}")
                continue
            ws = wb['ChlA']
            avg_col = None
            for col in range(1, ws.max_column+1):
                if ws.cell(row=1, column=col).value == 'Average':
                    avg_col = col
                    break
            if avg_col is None:
                logging.info("Column 'Average' not found.")
                continue
            avg_data = [ws.cell(row=r, column=avg_col).value for r in range(2, ws.max_row+1)]
            grid_file = find_file(os.path.join(os.path.dirname(source_file).replace("_result", "_grid")),
                                   rf"grid_.*{re.escape(indicator)}.*\.xlsx")
            _, grid_shape = read_points_from_xlsx_openpyxl(grid_file) if grid_file else (None, (20,))
            grid_rows = grid_shape[0] if grid_shape[0] is not None else 20
            wb_out = Workbook()
            ws1 = wb_out.active
            ws1.title = 'Sheet1'
            for r in range(1, grid_rows+1):
                ws1.cell(row=r, column=2, value=r).number_format = "0.0000"
            col_index = 3
            for i in range(0, len(avg_data), grid_rows):
                for j in range(grid_rows):
                    if i+j < len(avg_data):
                        ws1.cell(row=j+1, column=col_index, value=avg_data[i+j]).number_format = "0.0000"
                col_index += 1
            ws2 = wb_out.create_sheet(title='Sheet2')
            for col in range(3, col_index):
                col_letter = get_column_letter(col)
                for row in range(1, grid_rows+1):
                    ws2.cell(row=row, column=col, value=f"=Sheet1!{col_letter}{row}/AVERAGE(Sheet1!{col_letter}$1:{col_letter}$" + f"{grid_rows})")
            dummy = wb_out.create_sheet(title='DummyCalc')
            dummy['A1'] = "=NOW()"
            dummy.sheet_state = 'hidden'
            set_calculation_properties(wb_out)
            intermediate_file = os.path.join(output_folder, f"chart_{file_indicator}_intermediate.xlsx")
            wb_out.save(intermediate_file)
            wb2 = load_workbook(intermediate_file)
            if 'Sheet2' in wb2.sheetnames:
                ws2_2 = wb2['Sheet2']
                ws3 = wb2.create_sheet(title='Sheet3')
                dest_row = 1
                for col in range(1, ws2_2.max_column+1):
                    for row in range(1, ws2_2.max_row+1):
                        val = ws2_2.cell(row=row, column=col).value
                        if val is not None:
                            ws3.cell(row=dest_row, column=2, value=val)
                            dest_row += 1
                set_calculation_properties(wb2)
                final_chart = os.path.join(output_folder, f"chart_{file_indicator}.xlsx")
                wb2.save(final_chart)
                logging.info(f"Part D: Chart file generated: {final_chart}")

def run_partE(output_folder, indicator):
    logging.info(f"Part E: Skipping export of XLSM and CSV for indicator {indicator}")

#############################################
# INTERACTIVE SCATTER & EVOLUTION (Simplified Streamlit Version)
#############################################
def streamlit_interactive_scatter(all_data):
    st.header("Interactive Scatter & Evolution")
    if not all_data:
        st.error("No indicator data available for interactive scatter.")
        return
    indicators = list(all_data.keys())
    current_indicator = st.selectbox("Select Indicator", indicators)
    data_dict = all_data[current_indicator]
    # Determine coordinate system based on available data
    if data_dict.get('geo_points'):
        xvals, yvals = data_dict['geo_points']
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:3857", always_xy=True)
        xvals_trans = [transformer.transform(lon, lat)[0] for lon, lat in zip(xvals, yvals)]
        yvals_trans = [transformer.transform(lon, lat)[1] for lon, lat in zip(xvals, yvals)]
        scatter_coords = np.column_stack((xvals_trans, yvals_trans))
        lon_min, lon_max, lat_min, lat_max = data_dict['geo_bounds']
        x_min, y_min = transformer.transform(lon_min, lat_min)
        x_max, y_max = transformer.transform(lon_max, lat_max)
        extent = [x_min, x_max, y_max, y_min]
    else:
        bg_img = Image.open(data_dict['background'])
        w, h = bg_img.size
        xvals, yvals = data_dict['points']
        xvals_scaled = np.array(xvals) * (w / ref_width)
        yvals_scaled = np.array(yvals) * (h / ref_height)
        scatter_coords = np.column_stack((xvals_scaled, yvals_scaled))
        extent = [0, w, h, 0]
    # Sidebar controls
    size = st.sidebar.slider("Point Size", 1, 30, 10)
    min_val = st.sidebar.slider("Min Value", 0.0, 1.0, 0.8)
    max_val = st.sidebar.slider("Max Value", 1.0, 2.0, 1.2)
    alpha = st.sidebar.slider("Alpha", 0.0, 1.0, 0.7)
    exp_val = st.sidebar.slider("Exponent", -10.0, 0.0, -2.0)
    sharp = st.sidebar.slider("Sharp", 0.1, 2.0, 1.0)
    # Compute colors based on indicator data
    arr = data_dict['data']
    num_points = scatter_coords.shape[0]
    if len(arr) == 0:
        st.warning("Data array is empty; substituting with zeros.")
        arr = np.zeros(num_points)
    elif len(arr) != num_points:
        st.warning("Mismatch between data array and grid points; adjusting array length.")
        if len(arr) < num_points:
            arr = np.pad(arr, (0, num_points - len(arr)), mode='edge')
        else:
            arr = arr[:num_points]
    new_sizes = np.full(num_points, size)
    new_colors = []
    for val in arr:
        dist = abs(val - min_val) / max(max_val - min_val, 1e-6)
        gray = int(255 * np.exp(exp_val * dist) * sharp)
        gray = max(min(gray, 255), 0)
        new_colors.append((gray/255, gray/255, gray/255))
    # Plot the scatter overlaying the background using matplotlib
    fig, ax = plt.subplots(figsize=(7, 7))
    bg = Image.open(data_dict['background'])
    ax.imshow(bg, extent=extent, origin='upper', aspect='auto')
    ax.scatter(scatter_coords[:, 0], scatter_coords[:, 1],
               s=new_sizes, c=new_colors, alpha=alpha)
    ax.set_title(f"Indicator: {current_indicator}")
    st.pyplot(fig)
    st.write("Interactive scatter plot updated based on sidebar parameters.")

#############################################
# MAIN FUNCTION (Streamlit Version)
#############################################
def main_streamlit():
    st.title("Streamlit Environment for Image Processing and Grid Analysis")
    working_dir = st.text_input("Working Directory", os.getcwd())
    if not os.path.isdir(working_dir):
        st.error("Invalid working directory.")
        return
    os.chdir(working_dir)
    mother_folders = [folder for folder in os.listdir(working_dir)
                      if os.path.isdir(os.path.join(working_dir, folder))]
    valid_folders = []
    for folder in mother_folders:
        candidate = os.path.join(working_dir, folder)
        grid_sub = os.path.join(candidate, f"{folder}_grid")
        jpgs_sub = os.path.join(candidate, f"{folder}_jpgs")
        result_sub = os.path.join(candidate, f"{folder}_result")
        if os.path.isdir(grid_sub) and os.path.isdir(jpgs_sub):
            os.makedirs(result_sub, exist_ok=True)
            valid_folders.append(folder)
    if not valid_folders:
        st.error("No valid mother folders found with required subfolders.")
        return
    st.sidebar.subheader("Select Indicator")
    indicator = st.sidebar.selectbox("Indicator", valid_folders)
    mother_dir = os.path.join(working_dir, indicator)
    grid_dir = os.path.join(mother_dir, f"{indicator}_grid")
    jpgs_dir = os.path.join(mother_dir, f"{indicator}_jpgs")
    result_dir = os.path.join(mother_dir, f"{indicator}_result")
    # Check for JPG files; if missing, try converting a GIF
    jpgs = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
    if not jpgs:
        gif_files = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.gif')]
        if gif_files:
            gif_path = os.path.join(jpgs_dir, gif_files[0])
            st.write(f"No JPG found in {jpgs_dir}. Converting GIF: {gif_path}")
            gif_to_frames(gif_path, jpgs_dir)
    # Check for grid file; if not found, ask user to create one interactively
    if not find_file(grid_dir, r"grid_.*\.xlsx"):
        create = st.radio(f"No grid file found for indicator '{indicator}'. Do you want to create one interactively?", options=["Yes", "No"])
        if create == "Yes":
            grid_path = streamlit_create_grid_on_map(indicator, jpgs_dir)
            if grid_path:
                os.makedirs(grid_dir, exist_ok=True)
                os.replace(grid_path, os.path.join(grid_dir, os.path.basename(grid_path)))
            else:
                st.error(f"Grid creation cancelled for indicator '{indicator}'.")
                return
        else:
            st.error(f"No grid file for indicator '{indicator}' and user chose not to create one.")
            return
    # Process indicator if not complete
    if is_indicator_complete(result_dir, indicator):
        st.info(f"Indicator '{indicator}' is already complete. Skipping re-run.")
    else:
        st.info(f"Processing indicator: {indicator}")
        if not run_partA(grid_dir, jpgs_dir, result_dir, indicator):
            st.error(f"Indicator {indicator}: Part A failed. Skipping.")
            return
        run_partC(result_dir, indicator)
        run_partD(result_dir, indicator)
        run_partE(result_dir, indicator)
        chart_path = os.path.join(result_dir, f"chart_{indicator}.xlsx")
        force_recalc_with_excel(chart_path)
    # Prepare data for interactive scatter
    grid_pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
    chart_pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
    grid_file = find_file(grid_dir, grid_pattern)
    chart_file = find_file(result_dir, chart_pattern)
    if not grid_file or not chart_file:
        st.warning(f"Missing grid or chart file for indicator '{indicator}'.")
        return
    points, _ = read_points_from_xlsx_openpyxl(grid_file)
    if not points:
        st.warning(f"No points in grid file for indicator '{indicator}'.")
        return
    x_vals, y_vals = zip(*points)
    bounds = None
    kml_file = find_file(mother_dir, r".*\.kml")
    if kml_file:
        bounds = parse_kml_bounds(kml_file)
    else:
        txt_file = find_file(mother_dir, r".*\.txt")
        if txt_file:
            bounds = parse_txt_bounds(txt_file)
    geo_points = None
    geo_bounds = None
    if bounds:
        lon_min, lon_max, lat_min, lat_max = bounds
        geo_bounds = [lon_min, lon_max, lat_min, lat_max]
        geo_x_vals = [lon_min + (xx/ref_width)*(lon_max - lon_min) for xx in x_vals]
        geo_y_vals = [lat_min + (yy/ref_height)*(lat_max - lat_min) for yy in y_vals]
        geo_points = (geo_x_vals, geo_y_vals)
    try:
        wb = load_workbook(chart_file, data_only=True)
        if 'Sheet3' not in wb.sheetnames:
            st.error(f"Sheet3 missing in {chart_file}.")
            return
        ws = wb['Sheet3']
        data_array = [float(row[0]) for row in ws.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True)
                      if isinstance(row[0], (int, float))]
        wb.close()
    except Exception as e:
        st.error(f"Error reading data from {chart_file}: {e}")
        return
    frame0 = os.path.join(jpgs_dir, 'frame_0.jpg')
    if not os.path.isfile(frame0):
        cand = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
        if cand:
            frame0 = os.path.join(jpgs_dir, cand[0])
        else:
            st.error(f"No background image in {jpgs_dir} for indicator '{indicator}'.")
            return
    results_file = os.path.join(result_dir, f"results_{indicator}.xlsx")
    all_data = {
        indicator: {
            'points': (list(x_vals), list(y_vals)),
            'geo_points': geo_points,
            'geo_bounds': geo_bounds,
            'data': np.array(data_array),
            'background': frame0,
            'results_file': results_file
        }
    }
    streamlit_interactive_scatter(all_data)

if __name__ == '__main__':
    freeze_support()
    main_streamlit()
