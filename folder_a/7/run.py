import sys
print("Python executable:", sys.executable)
print("sys.path:", sys.path)

import os
working_dir = os.getcwd()

# Define reference dimensions for the grid (originally 512×512)
ref_width = 512
ref_height = 512

import re
import logging
import time
import numpy as np
import pandas as pd
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from multiprocessing import Pool, cpu_count, freeze_support
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button, RadioButtons
from tkinter import Tk, simpledialog, messagebox
import win32com.client as win32
import xml.etree.ElementTree as ET
import contextily as ctx
from pyproj import Transformer
from matplotlib.path import Path as MplPath
from pathlib import Path
import json

#############################################
# GIF CONVERSION HELPER
#############################################

def gif_to_frames(gif_path, output_folder):
    """Convert a GIF into individual JPEG frames saved in output_folder."""
    with Image.open(gif_path) as img:
        i = 0
        while True:
            try:
                img.seek(i)
                frame = img.convert('RGB')
                frame.save(os.path.join(output_folder, f'frame_{i}.jpg'), 'JPEG')
                i += 1
            except EOFError:
                break
    print("GIF has been processed. Frames saved as JPGs in:", output_folder)

#############################################
# HELPER FUNCTIONS
#############################################

def find_file(folder, pattern):
    """Return the first file in folder matching the regex pattern (case-insensitive)."""
    regex = re.compile(pattern, re.IGNORECASE)
    for fname in os.listdir(folder):
        if regex.search(fname):
            return os.path.join(folder, fname)
    return None

def parse_kml_bounds(kml_path):
    """Extract (lon_min, lon_max, lat_min, lat_max) from a KML file."""
    try:
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        coords_text = root.find('.//kml:coordinates', ns).text.strip()
        coords = coords_text.split()
        unique_coords = []
        for coord in coords:
            parts = coord.split(',')
            if len(parts) >= 2:
                lon, lat = float(parts[0]), float(parts[1])
                if (lon, lat) not in unique_coords:
                    unique_coords.append((lon, lat))
        if len(unique_coords) < 4:
            return None
        return (unique_coords[0][0], unique_coords[1][0],
                unique_coords[3][1], unique_coords[0][1])
    except Exception as e:
        logging.error(f"Error parsing KML file {kml_path}: {e}")
        return None

def parse_txt_bounds(txt_path):
    """
    Extract (lon_min, lon_max, lat_min, lat_max) from a TXT file containing JSON data.
    Expected format: {"type":"Polygon","coordinates":[[[lon,lat],[lon,lat], ...]]}
    """
    try:
        with open(txt_path, 'r', encoding='utf-8-sig') as f:
            content = f.read().strip()
            if not content:
                logging.error(f"File {txt_path} is empty!")
                return None
            data = json.loads(content)
        if data.get("type") == "Polygon" and "coordinates" in data:
            coords = data["coordinates"][0]
            lons = [pt[0] for pt in coords]
            lats = [pt[1] for pt in coords]
            return min(lons), max(lons), min(lats), max(lats)
        else:
            logging.error(f"JSON in {txt_path} does not have expected keys.")
            return None
    except Exception as e:
        logging.error(f"Error parsing txt file {txt_path}: {e}")
        return None

def set_calculation_properties(workbook):
    try:
        cp = workbook.calculation_properties
        cp.calculationMode = "auto"
        cp.calcCompleted = False
        cp.calcOnSave = True
        cp.fullCalcOnLoad = True
        cp.forceFullCalc = True
    except AttributeError:
        logging.warning("Calculation properties not supported in this openpyxl version.")

def force_recalc_with_excel(final_file):
    if not os.path.isfile(final_file):
        logging.error(f"Cannot recalc: File not found: {final_file}")
        return
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(final_file, UpdateLinks=False)
        wb.Application.CalculateFullRebuild()
        while wb.Application.CalculationState != 0:
            time.sleep(0.5)
        wb.Save()
        wb.Close(False)
        excel.Quit()
        logging.info(f"Excel recalculation forced and file saved: {final_file}")
    except Exception as e:
        logging.error(f"Error during Excel recalculation: {e}")

def is_indicator_complete(result_folder, indicator):
    pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
    return bool(find_file(result_folder, pattern))

#############################################
# READ RGB VALUES HELPER FUNCTION
#############################################

def read_rgb_values_at_point(image, point):
    try:
        pixel = image.getpixel((int(point[0]), int(point[1])))
        if isinstance(pixel, (int, float)):
            return (pixel, pixel, pixel)
        return pixel[:3]
    except Exception as e:
        logging.error(f"Error reading pixel at {point}: {e}")
        return None

#############################################
# READING GRID AND RESULTS
#############################################

def read_points_from_xlsx_openpyxl(file_path):
    points = []
    grid_shape = (None, None)
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        points = [(float(row[0]), float(row[1])) for row in sheet.iter_rows(min_row=2, values_only=True)
                  if row[0] is not None and row[1] is not None]
        n_rows = sheet['D2'].value
        n_cols = sheet['E2'].value
        grid_shape = (int(n_rows) if n_rows is not None else None,
                      int(n_cols) if n_cols is not None else None)
    except Exception as e:
        logging.error(f"Error reading grid file {file_path}: {e}")
    finally:
        wb.close()
    return points, grid_shape

def get_time_series_for_points(results_file, indices):
    wb = load_workbook(results_file, data_only=True)
    time_series = {idx: [] for idx in indices}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx in indices:
            try:
                val = float(rows[idx][5])
            except (IndexError, TypeError):
                val = None
            time_series[idx].append(val)
    wb.close()
    return time_series

#############################################
# OCR FUNCTIONS
#############################################

import pytesseract
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def pil_multi_threshold_preprocess(roi, thresholds=[80,120,160,200], invert=True):
    gray = roi.convert("L")
    for thr in thresholds:
        bin_img = gray.point(lambda p: 255 if p > thr else 0)
        yield bin_img
        if invert:
            yield ImageOps.invert(bin_img)

def pil_attempt_ocr(preprocessed_images, config):
    for img in preprocessed_images:
        text = pytesseract.image_to_string(img, config=config).strip()
        found = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        if found:
            dt_str = found[0]
            try:
                return datetime.strptime(dt_str, "%Y-%m-%d")
            except:
                continue
    return None

def extract_date(image_path):
    config = "--psm 7 -c tessedit_char_whitelist=0123456789-"
    try:
        img = Image.open(str(image_path))
    except Exception as e:
        print(f"Error: Unable to load image at {image_path}: {e}")
        return None
    width, height = img.size
    roi_dyn = img.crop((int(0.75*width), 0, width, int(0.15*height)))
    roi_dyn = roi_dyn.resize((roi_dyn.width*2, roi_dyn.height*2), Image.LANCZOS)
    dyn_images = pil_multi_threshold_preprocess(roi_dyn, thresholds=[80,120,160,200], invert=True)
    date_dyn = pil_attempt_ocr(dyn_images, config)
    if date_dyn:
        print(f"Dynamic OCR from {Path(image_path).name}: recognized {date_dyn.strftime('%Y-%m-%d')}")
        return date_dyn
    else:
        print(f"Dynamic OCR failed for {Path(image_path).name}, attempting fixed cropping.")
    roi_fixed = img.crop((width-160, 0, width, 60))
    roi_fixed = roi_fixed.resize((roi_fixed.width*2, roi_fixed.height*2), Image.LANCZOS)
    fixed_images = pil_multi_threshold_preprocess(roi_fixed, thresholds=[80,120,160,200], invert=True)
    date_fixed = pil_attempt_ocr(fixed_images, config)
    if date_fixed:
        print(f"Fixed OCR from {Path(image_path).name}: recognized {date_fixed.strftime('%Y-%m-%d')}")
        return date_fixed
    else:
        print(f"Date not found in {Path(image_path).name}")
        return None

def extract_number(filename):
    number = re.findall(r'\d+$', filename.stem)
    return int(number[0]) if number else 0

#############################################
# OCR RESULTS SAVING & READING
#############################################

def save_ocr_results(jpgs_dir):
    folder_path = Path(jpgs_dir)
    data = []
    for filename in sorted(folder_path.iterdir()):
        if filename.suffix.lower() in ('.png','.jpg','.jpeg'):
            frame_num = extract_number(filename)
            dt = extract_date(filename)
            data.append((frame_num, dt))
    data.sort(key=lambda x: x[0])
    wb_ocr = Workbook()
    ws = wb_ocr.active
    ws.title = "OCR_Dates"
    ws.append(["FrameNumber", "RecognizedDate"])
    for (fnum, dt) in data:
        dt_str = dt.strftime("%Y-%m-%d") if dt else ""
        ws.append([fnum, dt_str])
    out_path = folder_path / "ocr_results.xlsx"
    wb_ocr.save(str(out_path))
    print(f"OCR results saved to {out_path}")
    return out_path

def read_ocr_dates(jpgs_dir):
    folder_path = Path(jpgs_dir)
    ocr_file = folder_path / "ocr_results.xlsx"
    if not ocr_file.exists():
        ocr_file = save_ocr_results(jpgs_dir)
    wb = load_workbook(str(ocr_file), data_only=True)
    ws = wb.active
    frame_dates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        frame_num = row[0]
        date_str = row[1]
        dt = None
        if date_str:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
            except:
                pass
        frame_dates[frame_num] = dt
    wb.close()
    if frame_dates:
        max_frame = max(frame_dates.keys())
        out = [None]*(max_frame+1)
        for i in range(max_frame+1):
            out[i] = frame_dates.get(i, None)
        return out
    else:
        return []

#############################################
# GRID CREATION FUNCTIONS
#############################################

def create_grid_on_map(indicator, jpgs_dir):
    """
    MODIFIED: Both polygon and origin modes now use top-left origin in the preview.
    """
    root = Tk()
    root.withdraw()
    use_polygon = messagebox.askyesno(
        "Grid Creation Mode",
        "Do you want to define the grid by drawing a polygon?\n(If No, you'll set an origin with sliders.)",
        parent=root
    )
    grid_pts = []
    if use_polygon:
        messagebox.showinfo("Polygon Input", "Click on the image to define the polygon vertices. Press Enter when done.", parent=root)
        fig_poly, ax_poly = plt.subplots()
        candidates = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
        if not candidates:
            messagebox.showerror("Error", "No background image found in " + jpgs_dir, parent=root)
            root.destroy()
            return None
        bg_path = os.path.join(jpgs_dir, "frame_0.jpg") if os.path.isfile(os.path.join(jpgs_dir, "frame_0.jpg")) else os.path.join(jpgs_dir, candidates[0])
        bg_img = Image.open(bg_path)
        # Use top-left:
        ax_poly.imshow(bg_img, extent=[0, ref_width, ref_height, 0], origin='upper')
        ax_poly.set_title("Draw polygon vertices then press Enter")
        original_polygon = plt.ginput(n=-1, timeout=0)
        plt.close(fig_poly)
        if len(original_polygon) < 3:
            messagebox.showerror("Error", "A polygon must have at least 3 points.", parent=root)
            root.destroy()
            return None
        poly_arr = np.array(original_polygon)
        centroid = poly_arr.mean(axis=0)
        def rotate_points(points, angle_deg, center):
            angle_rad = np.radians(angle_deg)
            cos_a, sin_a = np.cos(angle_rad), np.sin(angle_rad)
            return [[(pt[0]-center[0])*cos_a - (pt[1]-center[1])*sin_a + center[0],
                     (pt[0]-center[0])*sin_a + (pt[1]-center[1])*cos_a + center[1]] for pt in points]
        spacing_h_init = float(simpledialog.askstring("Grid Spacing", "Enter initial horizontal grid spacing:", parent=root))
        spacing_v_init = spacing_h_init
        grid_rot_init = 0.0
        poly_rot_init = 0.0
        def generate_grid(h_spacing, v_spacing, poly_rot, grid_rot):
            adjusted_poly = rotate_points(original_polygon, poly_rot, centroid)
            poly_path = MplPath(adjusted_poly)
            xs, ys = zip(*adjusted_poly)
            bx_min, bx_max = min(xs), max(xs)
            by_min, by_max = min(ys), max(ys)
            pts = []
            for yy in np.arange(by_min, by_max + v_spacing, v_spacing):
                for xx in np.arange(bx_min, bx_max + h_spacing, h_spacing):
                    pts.append([xx, yy])
            grid_center = [(bx_min+bx_max)/2, (by_min+by_max)/2]
            pts_rot = rotate_points(pts, grid_rot, grid_center)
            valid_pts = [pt for pt in pts_rot if poly_path.contains_point((pt[0], pt[1]))]
            valid_pts.sort(key=lambda pt: (pt[1], pt[0]))
            return adjusted_poly, valid_pts
        fig_preview, ax_preview = plt.subplots()
        plt.subplots_adjust(bottom=0.35)
        ax_preview.imshow(bg_img, extent=[0, ref_width, ref_height, 0], origin='upper')
        preview_scatter = ax_preview.scatter([], [], color='red', s=10, label='Grid Points')
        poly_line, = ax_preview.plot([], [], color='blue', linestyle='--', linewidth=2, label='Polygon')
        ax_preview.set_title("Preview Grid.\nAdjust sliders then close window when satisfied.")
        ax_preview.legend()
        ax_hspacing = plt.axes([0.15, 0.25, 0.7, 0.03])
        ax_vspacing = plt.axes([0.15, 0.20, 0.7, 0.03])
        ax_gridrot = plt.axes([0.15, 0.15, 0.7, 0.03])
        ax_polyrot = plt.axes([0.15, 0.10, 0.7, 0.03])
        slider_hspacing = Slider(ax_hspacing, 'Horiz Spacing', 1, 100, valinit=spacing_h_init)
        slider_vspacing = Slider(ax_vspacing, 'Vert Spacing', 1, 100, valinit=spacing_v_init)
        slider_gridrot = Slider(ax_gridrot, 'Grid Rotation', -180, 180, valinit=grid_rot_init)
        slider_polyrot = Slider(ax_polyrot, 'Polygon Rotation', -180, 180, valinit=poly_rot_init)
        def update(val):
            h_spacing = slider_hspacing.val
            v_spacing = slider_vspacing.val
            gr = slider_gridrot.val
            pr = slider_polyrot.val
            adjusted_poly, new_grid = generate_grid(h_spacing, v_spacing, pr, gr)
            preview_scatter.set_offsets(np.array(new_grid) if new_grid else [])
            poly_arr2 = np.array(adjusted_poly)
            poly_line.set_data(np.append(poly_arr2[:,0], poly_arr2[0,0]),
                               np.append(poly_arr2[:,1], poly_arr2[0,1]))
            fig_preview.canvas.draw_idle()
        slider_hspacing.on_changed(update)
        slider_vspacing.on_changed(update)
        slider_gridrot.on_changed(update)
        slider_polyrot.on_changed(update)
        update(None)
        plt.show()
        _, grid_pts = generate_grid(slider_hspacing.val, slider_vspacing.val, slider_polyrot.val, slider_gridrot.val)
    else:
        logging.info("Using Origin Mode for grid creation.")
        fig_origin, ax_origin = plt.subplots()
        candidates = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
        if not candidates:
            messagebox.showerror("Error", "No background image found in " + jpgs_dir, parent=root)
            root.destroy()
            return None
        bg_path = os.path.join(jpgs_dir, "frame_0.jpg") if os.path.isfile(os.path.join(jpgs_dir, "frame_0.jpg")) else os.path.join(jpgs_dir, candidates[0])
        bg_img = Image.open(bg_path)
        ax_origin.imshow(bg_img, extent=[0, ref_width, ref_height, 0], origin='upper')
        ax_origin.set_title("Click to set grid origin (top-left)")
        origin = plt.ginput(1)
        if not origin:
            plt.close(fig_origin)
            root.destroy()
            return None
        origin = origin[0]
        logging.info(f"User selected origin: {origin}")
        plt.close(fig_origin)
        fig, ax = plt.subplots()
        ax.imshow(bg_img, extent=[0, ref_width, ref_height, 0], origin='upper')
        ax.set_title("Adjust grid parameters and click 'Save Grid'")
        init_spacing = 50
        init_rows = 10
        init_cols = 10
        init_angle = 0
        slider_ax_angle = plt.axes([0.25, 0.20, 0.65, 0.03])
        slider_angle = Slider(slider_ax_angle, 'Angle', -90, 90, valinit=init_angle)
        slider_ax_spacing = plt.axes([0.25, 0.15, 0.65, 0.03])
        slider_spacing = Slider(slider_ax_spacing, 'Spacing', 10, 200, valinit=init_spacing)
        slider_ax_rows = plt.axes([0.25, 0.10, 0.65, 0.03])
        slider_rows = Slider(slider_ax_rows, 'Rows', 1, 50, valinit=init_rows, valfmt='%0.0f')
        slider_ax_cols = plt.axes([0.25, 0.05, 0.65, 0.03])
        slider_cols = Slider(slider_ax_cols, 'Cols', 1, 50, valinit=init_cols, valfmt='%0.0f')
        grid_pts = []
        def update_grid(val):
            nonlocal grid_pts
            spacing = slider_spacing.val
            n_rows = int(slider_rows.val)
            n_cols = int(slider_cols.val)
            angle = slider_angle.val
            import math
            angle_rad = math.radians(angle)
            grid_pts = []
            for i in range(n_rows):
                for j in range(n_cols):
                    xx = j * spacing
                    yy = i * spacing
                    x_rot = xx * math.cos(angle_rad) - yy * math.sin(angle_rad)
                    y_rot = xx * math.sin(angle_rad) + yy * math.cos(angle_rad)
                    grid_pts.append([origin[0] + x_rot, origin[1] + y_rot])
            grid_pts.sort(key=lambda pt: (pt[1], pt[0]))
            for coll in ax.collections[:]:
                coll.remove()
            if grid_pts:
                xs, ys = zip(*grid_pts)
                ax.scatter(xs, ys, c='red', s=10)
            fig.canvas.draw_idle()
        slider_spacing.on_changed(update_grid)
        slider_rows.on_changed(update_grid)
        slider_cols.on_changed(update_grid)
        slider_angle.on_changed(update_grid)
        update_grid(None)
        ax_button = plt.axes([0.8, 0.9, 0.1, 0.05])
        button = Button(ax_button, 'Save Grid')
        grid_pts_final = None
        def save_grid(event):
            nonlocal grid_pts_final
            grid_pts_final = grid_pts
            plt.close(fig)
        button.on_clicked(save_grid)
        plt.show()
        if not grid_pts_final:
            root.destroy()
            return None
        grid_pts = grid_pts_final
    root.update_idletasks()
    try:
        root.destroy()
    except Exception:
        pass
    grid_pts.sort(key=lambda pt: (pt[1], pt[0]))
    wb = Workbook()
    ws = wb.active
    ws.title = "GridPoints"
    ws.append(["X", "Y"])
    for pt in grid_pts:
        ws.append(pt)
    if use_polygon:
        ys = [pt[1] for pt in grid_pts]
        unique_y = np.unique(np.round(ys, 2))
        n_rows = len(unique_y)
        n_cols = int(np.ceil(len(grid_pts) / n_rows))
    else:
        n_rows = int(slider_rows.val)
        n_cols = int(slider_cols.val)
    ws['D1'] = "Rows"
    ws['D2'] = n_rows
    ws['E1'] = "Cols"
    ws['E2'] = n_cols
    grid_filename = f"grid_{indicator}.xlsx"
    save_path = os.path.join(os.getcwd(), grid_filename)
    wb.save(save_path)
    logging.info(f"Grid file created: {save_path}")
    return save_path

#############################################
# PART A: Process Images Based on Grid
#############################################

def process_image(image_path, points):
    """
    Because we now define the preview with top-left,
    we simply scale: x_scaled = x*(img_width/ref_width), y_scaled = y*(img_height/ref_height).
    """
    img = Image.open(image_path).convert("RGB")
    img_width, img_height = img.size
    scaled_points = [
        (xx * img_width / ref_width,
         yy * img_height / ref_height)
        for xx, yy in points
    ]
    results = []
    for pt in scaled_points:
        rgb = read_rgb_values_at_point(img, pt)
        if rgb:
            r, g, b = rgb
            grayscale = 0.299*r + 0.587*g + 0.114*b
            results.append([pt[0], pt[1], round(r, 3), round(g, 3), round(b, 3), round(grayscale, 3)])
    img.close()
    return results

def write_results_to_excel(results, workbook_path):
    wb = Workbook()
    wb.remove(wb.active)
    for index, data in enumerate(results):
        sheet = wb.create_sheet(title=str(index))
        sheet.append(["X", "Y", "R", "G", "B", "Grayscale"])
        for row in data:
            sheet.append(row)
    set_calculation_properties(wb)
    wb.save(workbook_path)

def run_partA(input_grid_folder, input_images_folder, output_folder, indicator):
    pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
    grid_file = find_file(input_grid_folder, pattern)
    if not grid_file:
        logging.error(f"Grid file not found in {input_grid_folder} for indicator {indicator}")
        return False
    points, _ = read_points_from_xlsx_openpyxl(grid_file)
    if not points:
        logging.error(f"No points read from {grid_file}. Exiting Part A.")
        return False
    jpg_files = [
        os.path.join(input_images_folder, f)
        for f in os.listdir(input_images_folder)
        if f.lower().endswith('.jpg')
    ]
    if not jpg_files:
        logging.error(f"No .jpg files found in {input_images_folder}.")
        return False
    with Pool(cpu_count()) as pool:
        results = pool.starmap(process_image, [(jpg, points) for jpg in jpg_files])
    output_xlsx = os.path.join(output_folder, f"results_{indicator}.xlsx")
    write_results_to_excel(results, output_xlsx)
    logging.info(f"Part A completed. Wrote {output_xlsx}")
    return True

#############################################
# PART C
#############################################

def run_partC(output_folder, indicator):
    import pandas as pd
    for filename in os.listdir(output_folder):
        if filename.startswith("~$") or filename.startswith("chart_"):
            continue
        if filename.endswith(".xlsx"):
            input_file = os.path.join(output_folder, filename)
            match = re.search(r'_([A-Za-z0-9]+)\.xlsx$', filename)
            if not match:
                continue
            file_indicator = match.group(1)
            out_file = os.path.join(output_folder, f"total_results_{file_indicator}.xlsx")
            wb = load_workbook(input_file, read_only=True, data_only=True)
            data_frames = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = list(sheet.values)
                if not data:
                    continue
                data = data[1:]
                df = pd.DataFrame(data)
                try:
                    col_data = df.iloc[:, 3].dropna()
                    if not col_data.empty:
                        data_frames.append(col_data.reset_index(drop=True).to_frame(name=sheet_name))
                except IndexError:
                    logging.info(f"No Column 'D' in {sheet_name}")
            if data_frames:
                all_data_df = pd.concat(data_frames, axis=1)
                all_data_df['Average'] = all_data_df.mean(axis=1)
                with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                    all_data_df.to_excel(writer, sheet_name='ChlA', index=False)
                    set_calculation_properties(writer.book)
                logging.info(f"Part C: Wrote {out_file}")

#############################################
# PART D
#############################################

def run_partD(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("~$"):
            continue
        if filename.startswith("total_results_") and filename.endswith(".xlsx"):
            source_file = os.path.join(output_folder, filename)
            file_indicator = indicator
            wb = load_workbook(source_file, data_only=True)
            if 'ChlA' not in wb.sheetnames:
                logging.info(f"Sheet 'ChlA' not found in {filename}")
                continue
            ws = wb['ChlA']
            avg_col = None
            for col in range(1, ws.max_column+1):
                if ws.cell(row=1, column=col).value == 'Average':
                    avg_col = col
                    break
            if avg_col is None:
                logging.info("Column 'Average' not found.")
                continue
            avg_data = [ws.cell(row=r, column=avg_col).value for r in range(2, ws.max_row+1)]
            grid_file = find_file(os.path.join(os.path.dirname(source_file).replace("_result", "_grid")),
                                   rf"grid_.*{re.escape(indicator)}.*\.xlsx")
            _, grid_shape = read_points_from_xlsx_openpyxl(grid_file) if grid_file else (None, (20,))
            grid_rows = grid_shape[0] if grid_shape[0] is not None else 20
            wb_out = Workbook()
            ws1 = wb_out.active
            ws1.title = 'Sheet1'
            for r in range(1, grid_rows+1):
                ws1.cell(row=r, column=2, value=r).number_format = "0.0000"
            col_index = 3
            for i in range(0, len(avg_data), grid_rows):
                for j in range(grid_rows):
                    if i+j < len(avg_data):
                        ws1.cell(row=j+1, column=col_index, value=avg_data[i+j]).number_format = "0.0000"
                col_index += 1
            ws2 = wb_out.create_sheet(title='Sheet2')
            for col in range(3, col_index):
                col_letter = get_column_letter(col)
                for row in range(1, grid_rows+1):
                    ws2.cell(row=row, column=col, value=f"=Sheet1!{col_letter}{row}/AVERAGE(Sheet1!{col_letter}$1:{col_letter}$" + f"{grid_rows})")
            dummy = wb_out.create_sheet(title='DummyCalc')
            dummy['A1'] = "=NOW()"
            dummy.sheet_state = 'hidden'
            set_calculation_properties(wb_out)
            intermediate_file = os.path.join(output_folder, f"chart_{file_indicator}_intermediate.xlsx")
            wb_out.save(intermediate_file)
            wb2 = load_workbook(intermediate_file)
            if 'Sheet2' in wb2.sheetnames:
                ws2_2 = wb2['Sheet2']
                ws3 = wb2.create_sheet(title='Sheet3')
                dest_row = 1
                for col in range(1, ws2_2.max_column+1):
                    for row in range(1, ws2_2.max_row+1):
                        val = ws2_2.cell(row=row, column=col).value
                        if val is not None:
                            ws3.cell(row=dest_row, column=2, value=val)
                            dest_row += 1
                set_calculation_properties(wb2)
                final_chart = os.path.join(output_folder, f"chart_{file_indicator}.xlsx")
                wb2.save(final_chart)
                logging.info(f"Part D: Chart file generated: {final_chart}")

def run_partE(output_folder, indicator):
    logging.info(f"Part E: Skipping export of XLSM and CSV for indicator {indicator}")

#############################################
# INTERACTIVE SCATTER & EVOLUTION
#############################################

def run_interactive_scatter_for_all_indicators(all_data):
    """
    Now that both preview and final use top-left,
    we simply scale:
      x_scaled = x * (w/ref_width) and y_scaled = y * (h/ref_height)
    """
    if not all_data:
        logging.error("No indicator data available for interactive scatter. Exiting.")
        return

    import matplotlib.pyplot as plt
    from PIL import Image
    import matplotlib.dates as mdates
    from pyproj import Transformer

    indicators = list(all_data.keys())
    current_indicator = indicators[0]
    indicator_limits = {}

    def get_current_data():
        return all_data[current_indicator]

    def update_scatter():
        d = get_current_data()
        if d.get('geo_points'):
            xvals, yvals = d['geo_points']
            transformer = Transformer.from_crs("EPSG:4326", "EPSG:3857", always_xy=True)
            xvals_trans = [transformer.transform(lon, lat)[0] for lon, lat in zip(xvals, yvals)]
            yvals_trans = [transformer.transform(lon, lat)[1] for lon, lat in zip(xvals, yvals)]
            scatter_coords = np.column_stack((xvals_trans, yvals_trans))
            lon_min, lon_max, lat_min, lat_max = d['geo_bounds']
            x_min, y_min = transformer.transform(lon_min, lat_min)
            x_max, y_max = transformer.transform(lon_max, lat_max)
            default_extent = [x_min, x_max, y_max, y_min]
        else:
            bg_img = Image.open(d['background'])
            w, h = bg_img.size
            xvals, yvals = d['points']
            xvals_scaled = np.array(xvals) * (w / ref_width)
            yvals_scaled = np.array(yvals) * (h / ref_height)
            scatter_coords = np.column_stack((xvals_scaled, yvals_scaled))
            default_extent = [0, w, h, 0]

        if current_indicator in indicator_limits:
            xlim, ylim = indicator_limits[current_indicator]
            ax1.set_xlim(xlim)
            ax1.set_ylim(ylim)
            ax2.set_xlim(xlim)
            ax2.set_ylim(ylim)
        else:
            ax1.set_xlim(default_extent[0], default_extent[1])
            ax1.set_ylim(default_extent[2], default_extent[3])
            ax2.set_xlim(default_extent[0], default_extent[1])
            ax2.set_ylim(default_extent[2], default_extent[3])

        for im in ax1.images + ax2.images:
            new_bg = Image.open(d['background'])
            im.set_data(new_bg)
            im.set_extent(default_extent)

        scatter1.set_offsets(scatter_coords)
        scatter2.set_offsets(scatter_coords)

        arr = d['data']
        new_sizes = np.full(len(arr), slider_size.val)
        new_colors = []
        for val in arr:
            dist = abs(val - slider_min.val) / max(slider_max.val - slider_min.val, 1e-6)
            gray = int(255 * np.exp(slider_exp.val * dist) * slider_sharp.val)
            gray = max(min(gray, 255), 0)
            new_colors.append((gray/255, gray/255, gray/255))
        scatter1.set_sizes(new_sizes)
        scatter1.set_facecolors(new_colors)
        scatter1.set_alpha(slider_alpha1.val)
        scatter2.set_sizes(new_sizes)
        scatter2.set_facecolors(new_colors)
        scatter2.set_alpha(slider_alpha2.val)

        ax1.set_title(f"Indicator: {current_indicator} (Left Plot)")
        ax2.set_title(f"Indicator: {current_indicator} (Right Plot)")
        fig.canvas.draw_idle()

    def on_indicator_label_clicked(label):
        nonlocal current_indicator
        indicator_limits[current_indicator] = (ax1.get_xlim(), ax1.get_ylim())
        current_indicator = label
        update_scatter()

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 7))
    plt.subplots_adjust(left=0.25, bottom=0.35, top=0.95)

    def on_xlim_ylim_changed(event_ax):
        indicator_limits[current_indicator] = (ax1.get_xlim(), ax1.get_ylim())
        ax2.set_xlim(ax1.get_xlim())
        ax2.set_ylim(ax1.get_ylim())
        fig.canvas.draw_idle()

    ax1.callbacks.connect('xlim_changed', on_xlim_ylim_changed)
    ax1.callbacks.connect('ylim_changed', on_xlim_ylim_changed)

    initial_data = get_current_data()
    if initial_data.get('geo_bounds'):
        transformer = Transformer.from_crs("EPSG:4326", "EPSG:3857", always_xy=True)
        lon_min, lon_max, lat_min, lat_max = initial_data['geo_bounds']
        x_min, y_min = transformer.transform(lon_min, lat_min)
        x_max, y_max = transformer.transform(lon_max, lat_max)
        default_extent = [x_min, x_max, y_max, y_min]
    else:
        bg_img = Image.open(initial_data['background'])
        w, h = bg_img.size
        default_extent = [0, w, h, 0]
    ax1.imshow(Image.open(initial_data['background']), extent=default_extent, origin='upper', aspect='auto')
    ax2.imshow(Image.open(initial_data['background']), extent=default_extent, origin='upper', aspect='auto')

    scatter1 = ax1.scatter([], [], c='white')
    scatter2 = ax2.scatter([], [], c='white')

    ax_size = plt.axes([0.25, 0.2, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_min_slider = plt.axes([0.25, 0.17, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_max_slider = plt.axes([0.25, 0.14, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_alpha1 = plt.axes([0.25, 0.11, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_alpha2 = plt.axes([0.25, 0.08, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_exp = plt.axes([0.25, 0.05, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_sharp = plt.axes([0.25, 0.02, 0.65, 0.02], facecolor='lightgoldenrodyellow')

    slider_size = Slider(ax_size, 'Size', 1, 30, valinit=29.9)
    slider_min = Slider(ax_min_slider, 'Min', 0.0, 1.0, valinit=0.838)
    slider_max = Slider(ax_max_slider, 'Max', 1.0, 2.0, valinit=1.252)
    slider_alpha1 = Slider(ax_alpha1, 'Alpha1', 0.0, 1.0, valinit=0.687)
    slider_alpha2 = Slider(ax_alpha2, 'Alpha2', 0.0, 1.0, valinit=0.05)
    slider_exp = Slider(ax_exp, 'Exponent', -10, 0, valinit=-2.02)
    slider_sharp = Slider(ax_sharp, 'Sharp', 0.1, 2.0, valinit=1.192)

    ax_radio = plt.axes([0.05, 0.35, 0.15, 0.5], facecolor='lightgoldenrodyellow')
    indicators = list(all_data.keys()) if all_data else []
    radio_buttons = RadioButtons(ax_radio, indicators, active=0)
    radio_buttons.on_clicked(on_indicator_label_clicked)

    def slider_updated(val):
        update_scatter()
    slider_size.on_changed(slider_updated)
    slider_min.on_changed(slider_updated)
    slider_max.on_changed(slider_updated)
    slider_alpha1.on_changed(slider_updated)
    slider_alpha2.on_changed(slider_updated)
    slider_exp.on_changed(slider_updated)
    slider_sharp.on_changed(slider_updated)

    update_scatter()

    ax_line = plt.axes([0.05, 0.05, 0.15, 0.05])
    btn_line = Button(ax_line, 'Show Evolution')
    def show_evolution(event):
        line_pts = plt.ginput(2, timeout=30)
        if len(line_pts) != 2:
            print("Line not defined properly.")
            return
        (lx1, ly1), (lx2, ly2) = line_pts
        ax1.plot([lx1, lx2], [ly1, ly2], color='cyan', linewidth=2)
        fig.canvas.draw_idle()
        print(f"Line drawn: ({lx1:.2f}, {ly1:.2f}) to ({lx2:.2f}, {ly2:.2f})")
        d = get_current_data()
        points_len = len(d['points'][0])
        selected_indices = [0, points_len - 1] if points_len >= 2 else [0]
        print("Selected grid point indices:", selected_indices)
        ts_dict = get_time_series_for_points(d['results_file'], selected_indices)
        num_points = len(selected_indices)
        max_frames = max(len(series) for series in ts_dict.values())
        combined_data = np.zeros((num_points, max_frames, 3), dtype=np.float32)
        for row_i, (pt_idx, series) in enumerate(ts_dict.items()):
            for frame_i, val in enumerate(series):
                if val is None:
                    val = 0
                g = max(0, min(255, val))
                combined_data[row_i, frame_i, :] = [g/255.0, g/255.0, g/255.0]
        fig2, (ax_top, ax_bottom) = plt.subplots(2, 1, figsize=(12, 8), sharex=False)
        fig2.subplots_adjust(hspace=0.3)
        ax_top.imshow(combined_data, aspect='auto')
        ax_top.set_title("Grayscale Variation (vs Frames) for Selected Grid Points")
        ax_top.set_ylabel("Grid Point Index")
        ax_top.set_xlabel("Frame Index")
        ax_top.set_yticks(range(num_points))
        ax_top.set_yticklabels([f"Point {idx}" for idx in selected_indices])
        for row_i, (pt_idx, series) in enumerate(ts_dict.items()):
            norm_series = [max(0, min(255, v if v is not None else 0))/255.0 for v in series]
            line_y = [row_i + v for v in norm_series]
            line_x = list(range(len(series)))
            ax_top.plot(line_x, line_y, color='red', linewidth=2)
        ax_bottom.set_title("Grayscale Variation (vs Date) for Selected Grid Points")
        ax_bottom.set_ylabel("Grayscale Value")
        ax_bottom.set_xlabel("Date")
        guess_jpgs = None
        if '_result' in os.path.dirname(d['results_file']):
            guess_jpgs = os.path.dirname(d['results_file']).replace('_result', '_jpgs')
        frame_dates = []
        if guess_jpgs and os.path.isdir(guess_jpgs):
            frame_dates = read_ocr_dates(guess_jpgs)
        for pt_idx, series in ts_dict.items():
            seg_x = []
            seg_y = []
            for i in range(len(series)):
                dt = frame_dates[i] if i < len(frame_dates) else None
                if dt is not None and series[i] is not None:
                    seg_x.append(dt)
                    seg_y.append(series[i])
                else:
                    if seg_x and seg_y:
                        ax_bottom.plot(seg_x, seg_y, color='blue', linewidth=2, label=f"Pt {pt_idx}")
                    seg_x = []
                    seg_y = []
            if seg_x and seg_y:
                ax_bottom.plot(seg_x, seg_y, color='blue', linewidth=2, label=f"Pt {pt_idx}")
        ax_bottom.xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        fig2.autofmt_xdate()
        out_table = []
        for i in range(max_frames):
            date_val = frame_dates[i].strftime("%Y-%m-%d") if i < len(frame_dates) and frame_dates[i] is not None else str(i)
            row = [date_val]
            for pt_idx in selected_indices:
                val = ts_dict[pt_idx][i] if i < len(ts_dict[pt_idx]) else None
                row.append(val)
            out_table.append(row)
        wb_evo = Workbook()
        ws_evo = wb_evo.active
        ws_evo.title = "Evolution Data"
        header = ["Date/Frame"] + [f"Pt {idx}" for idx in selected_indices]
        ws_evo.append(header)
        for row in out_table:
            ws_evo.append(row)
        evo_file = os.path.join(guess_jpgs, "evolution_data.xlsx") if guess_jpgs else os.path.join(working_dir, "evolution_data.xlsx")
        wb_evo.save(evo_file)
        print(f"Evolution data saved to {evo_file}")
        out_fig = os.path.join(working_dir, "line_evolution.png")
        plt.savefig(out_fig, dpi=150)
        print(f"Saved evolution figure to {out_fig}")
        plt.show()

    btn_line.on_clicked(show_evolution)
    plt.show()

#############################################
# MAIN FUNCTION
#############################################

def main_option2_single_window():
    mother_folders = [folder for folder in os.listdir(working_dir) if os.path.isdir(os.path.join(working_dir, folder))]
    valid_folders = []
    for folder in mother_folders:
        candidate = os.path.join(working_dir, folder)
        grid_sub = os.path.join(candidate, f"{folder}_grid")
        jpgs_sub = os.path.join(candidate, f"{folder}_jpgs")
        result_sub = os.path.join(candidate, f"{folder}_result")
        if os.path.isdir(grid_sub) and os.path.isdir(jpgs_sub):
            os.makedirs(result_sub, exist_ok=True)
            valid_folders.append(folder)
    if not valid_folders:
        logging.error("No valid mother folders found with required subfolders.")
        return

    all_data = {}
    for indicator in valid_folders:
        mother_dir = os.path.join(working_dir, indicator)
        grid_dir = os.path.join(mother_dir, f"{indicator}_grid")
        jpgs_dir = os.path.join(mother_dir, f"{indicator}_jpgs")
        result_dir = os.path.join(mother_dir, f"{indicator}_result")

        jpgs = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
        if not jpgs:
            gif_files = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.gif')]
            if gif_files:
                gif_path = os.path.join(jpgs_dir, gif_files[0])
                print(f"No JPG found in {jpgs_dir}. Converting GIF: {gif_path}")
                gif_to_frames(gif_path, jpgs_dir)

        if not find_file(grid_dir, r"grid_.*\.xlsx"):
            create = messagebox.askyesno("Create Grid",
                f"No grid file found for indicator '{indicator}'.\nDo you want to create one interactively?")
            if create:
                grid_path = create_grid_on_map(indicator, jpgs_dir)
                if grid_path:
                    os.makedirs(grid_dir, exist_ok=True)
                    os.replace(grid_path, os.path.join(grid_dir, os.path.basename(grid_path)))
                else:
                    logging.error(f"Grid creation cancelled for indicator '{indicator}'. Skipping.")
                    continue
            else:
                logging.error(f"No grid file for indicator '{indicator}' and user chose not to create one. Skipping.")
                continue

        if is_indicator_complete(result_dir, indicator):
            logging.info(f"Indicator '{indicator}' is already complete. Skipping re-run.")
        else:
            logging.info(f"Processing indicator: {indicator}")
            if not run_partA(grid_dir, jpgs_dir, result_dir, indicator):
                logging.error(f"Indicator {indicator}: Part A failed. Skipping.")
                continue
            run_partC(result_dir, indicator)
            run_partD(result_dir, indicator)
            run_partE(result_dir, indicator)
            chart_path = os.path.join(result_dir, f"chart_{indicator}.xlsx")
            force_recalc_with_excel(chart_path)

        grid_pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
        chart_pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
        grid_file = find_file(grid_dir, grid_pattern)
        chart_file = find_file(result_dir, chart_pattern)
        if not grid_file or not chart_file:
            logging.warning(f"Missing grid or chart file for indicator '{indicator}'.")
            continue

        points, _ = read_points_from_xlsx_openpyxl(grid_file)
        if not points:
            logging.warning(f"No points in grid file for indicator '{indicator}'. Skipping.")
            continue

        x_vals, y_vals = zip(*points)
        bounds = None
        kml_file = find_file(mother_dir, r".*\.kml")
        if kml_file:
            bounds = parse_kml_bounds(kml_file)
        else:
            txt_file = find_file(mother_dir, r".*\.txt")
            if txt_file:
                bounds = parse_txt_bounds(txt_file)

        geo_points = None
        geo_bounds = None
        if bounds:
            lon_min, lon_max, lat_min, lat_max = bounds
            geo_bounds = [lon_min, lon_max, lat_min, lat_max]
            geo_x_vals = [lon_min + (xx/ref_width)*(lon_max - lon_min) for xx in x_vals]
            geo_y_vals = [lat_min + (yy/ref_height)*(lat_max - lat_min) for yy in y_vals]
            geo_points = (geo_x_vals, geo_y_vals)

        try:
            wb = load_workbook(chart_file, data_only=True)
            if 'Sheet3' not in wb.sheetnames:
                logging.error(f"Sheet3 missing in {chart_file}.")
                continue
            ws = wb['Sheet3']
            data_array = [float(row[0]) for row in ws.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True)
                          if isinstance(row[0], (int, float))]
            wb.close()
        except Exception as e:
            logging.error(f"Error reading data from {chart_file}: {e}")
            continue

        frame0 = os.path.join(jpgs_dir, 'frame_0.jpg')
        if not os.path.isfile(frame0):
            cand = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
            if cand:
                frame0 = os.path.join(jpgs_dir, cand[0])
            else:
                logging.error(f"No background image in {jpgs_dir}. Skipping {indicator}.")
                continue

        results_file = os.path.join(result_dir, f"results_{indicator}.xlsx")
        all_data[indicator] = {
            'points': (list(x_vals), list(y_vals)),
            'geo_points': geo_points,
            'geo_bounds': geo_bounds,
            'data': np.array(data_array),
            'background': frame0,
            'results_file': results_file
        }

    run_interactive_scatter_for_all_indicators(all_data)

if __name__ == '__main__':
    freeze_support()
    main_option2_single_window()