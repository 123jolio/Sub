import os
import re
import pytesseract
import cv2
import numpy as np
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

# Set the Tesseract executable path
pytesseract.pytesseract.tesseract_cmd = r'C:\Users\ilioumbas\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

def parse_date_string(text):
    """
    Attempt to find a date in the OCR text using a regex for YYYY-MM-DD.
    If found, return a datetime object. Also correct the year if it's > 2025.
    Otherwise, return None.
    """
    # Look for 'YYYY-MM-DD'
    candidates = re.findall(r'\d{4}-\d{2}-\d{2}', text)
    if not candidates:
        return None
    
    # Just use the first match if multiple
    date_str = candidates[0]
    try:
        year = int(date_str[:4])
        # If the year is suspiciously high (e.g., 2080), correct it
        if year > 2025:
            corrected_date_str = f"20{date_str[2:]}"  # e.g., 2080 -> 20 + 80 => 2080
            return datetime.strptime(corrected_date_str, '%Y-%m-%d')
        else:
            return datetime.strptime(date_str, '%Y-%m-%d')
    except ValueError:
        return None

def upscale_image(img, factor=2):
    """
    Upscale an image by a given factor to help Tesseract read small text.
    """
    return cv2.resize(img, None, fx=factor, fy=factor, interpolation=cv2.INTER_CUBIC)

def color_mask_for_white_text(img):
    """
    If the date text is mostly white/gray, we can apply a color-based mask in HSV space.
    Adjust the lower/upper bounds as needed for your images.
    """
    hsv = cv2.cvtColor(img, cv2.COLOR_BGR2HSV)
    # This range attempts to capture near-white or light gray tones
    lower = np.array([0, 0, 180])   # H=0, S=0, V=180
    upper = np.array([180, 50, 255])  # H=180, S=50, V=255
    mask = cv2.inRange(hsv, lower, upper)
    # Keep only masked region
    result = cv2.bitwise_and(img, img, mask=mask)
    return result

def preprocess_variants(img):
    """
    Generate multiple thresholded / morphological variants of the image
    to feed into Tesseract. Each variant might catch the text differently.
    """
    # Convert to grayscale
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    variants = []

    # 1) Simple Otsu threshold (inverse)
    _, th_inv = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)
    variants.append(th_inv)

    # 2) Simple Otsu threshold (normal)
    _, th = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)
    variants.append(th)

    # 3) Adaptive threshold
    adapt = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                  cv2.THRESH_BINARY, 31, 2)
    variants.append(adapt)

    # 4) Morphological close on the Otsu threshold
    kernel = np.ones((3, 3), np.uint8)
    th_close = cv2.morphologyEx(th, cv2.MORPH_CLOSE, kernel)
    variants.append(th_close)

    # Optionally add more morphological variants if needed
    # e.g., open, dilate, erode, etc.

    return variants

def attempt_ocr_on_roi(roi):
    """
    Perform multiple preprocessing steps on a given ROI and
    try Tesseract OCR on each. Return the first valid date found or None.
    """
    # Optional color-based mask if your text is consistently white
    # This can remove a lot of color noise in the background
    masked = color_mask_for_white_text(roi)

    # Upscale for better OCR
    masked_up = upscale_image(masked, factor=2)

    # Try multiple threshold / morphological variants
    for prep in preprocess_variants(masked_up):
        # Whitelist digits and dash to reduce confusion
        config = r'--oem 3 --psm 6 -c tessedit_char_whitelist=0123456789-'
        text = pytesseract.image_to_string(prep, config=config).strip()
        if text:
            print(f"[DEBUG] OCR text: '{text}'")
        date_obj = parse_date_string(text)
        if date_obj:
            return date_obj

    return None

def extract_date(image_path):
    """
    Attempt to extract a date from an image by OCR-ing the top-right corner,
    then falling back to a larger top strip if that fails.
    Returns a datetime object or None.
    """
    # Read image in a way that handles unicode paths
    img_data = np.fromfile(str(image_path), dtype=np.uint8)
    img = cv2.imdecode(img_data, cv2.IMREAD_COLOR)
    
    if img is None:
        print(f"[ERROR] Unable to load image at {image_path}")
        return None

    height, width, _ = img.shape
    
    # 1) First attempt: top-right corner (10% high, 25% wide)
    roi_top = 0
    roi_bottom = int(0.10 * height)
    roi_left = int(0.75 * width)
    roi_right = width
    corner_roi = img[roi_top:roi_bottom, roi_left:roi_right]

    date_obj = attempt_ocr_on_roi(corner_roi)
    if date_obj:
        return date_obj

    # 2) Second attempt: top strip (10% high, entire width)
    print(f"[INFO] Corner-based OCR failed, trying larger top strip for {image_path.name}")
    strip_roi = img[0:int(0.10 * height), 0:width]
    date_obj = attempt_ocr_on_roi(strip_roi)
    if date_obj:
        return date_obj

    print(f"[WARNING] Date not found in both corner and top strip for {image_path.name}")
    return None

def extract_number(filename):
    """
    Extract trailing digits from the filename stem (e.g. 'image_10' -> 10).
    If no digits found, return 0.
    """
    number = re.findall(r'\d+$', filename.stem)
    return int(number[0]) if number else 0

def main():
    # Set folder path to the 'jpgs' folder in the current working directory
    folder_path = Path(os.getcwd()) / 'jpgs'

    if not folder_path.exists():
        print(f"[ERROR] The folder {folder_path} does not exist.")
        return

    data = []
    # Process each image in sorted order
    for filename in sorted(folder_path.iterdir()):
        if filename.suffix.lower() in ('.png', '.jpg', '.jpeg'):
            print(f"[INFO] Processing file: {filename.name}")
            date_obj = extract_date(filename)
            file_number = extract_number(filename)
            if date_obj:
                data.append((file_number, date_obj.year, date_obj.month, date_obj.day))
            else:
                data.append((file_number, None, None, None))

    # Sort data (handle None gracefully)
    def sort_key(row):
        return tuple(x if x is not None else float('inf') for x in row)
    sorted_data = sorted(data, key=sort_key)

    # Save data to Excel
    output_path = folder_path / 'ocr.xlsx'
    wb = Workbook()
    ws = wb.active

    # Header
    ws.cell(row=1, column=1, value='File Number')
    ws.cell(row=1, column=2, value='Year')
    ws.cell(row=1, column=3, value='Month')
    ws.cell(row=1, column=4, value='Day')

    for row_idx, row_data in enumerate(sorted_data, start=2):
        for col_idx, cell_data in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=cell_data)

    wb.save(str(output_path))
    print(f"[INFO] Data saved to {output_path}")

if __name__ == "__main__":
    main()
