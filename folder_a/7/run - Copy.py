import os
import re
import logging
import time
import numpy as np
import pandas as pd
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from multiprocessing import Pool, cpu_count
import matplotlib.pyplot as plt
from matplotlib.widgets import Slider, Button, RadioButtons
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import win32com.client as win32
import xml.etree.ElementTree as ET  # for parsing KML

logging.basicConfig(level=logging.INFO, format=r'%(asctime)s - %(levelname)s - %(message)s')
working_dir = os.getcwd()

# ---------------------------------------------------------------------------
# Helper: Find a file matching a pattern (case-insensitive)
# ---------------------------------------------------------------------------
def find_file(folder, pattern):
    regex = re.compile(pattern, re.IGNORECASE)
    for fname in os.listdir(folder):
        if regex.search(fname):
            return os.path.join(folder, fname)
    return None

# ---------------------------------------------------------------------------
# KML Parsing Helper: Extract geographic bounds from a KML file
# ---------------------------------------------------------------------------
def parse_kml_bounds(kml_path):
    try:
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        coords_text = root.find('.//kml:coordinates', ns).text.strip()
        # Coordinates string is typically "lon,lat,0 lon,lat,0 ..."
        coords = coords_text.split()
        unique_coords = []
        for coord in coords:
            parts = coord.split(',')
            if len(parts) >= 2:
                lon, lat = float(parts[0]), float(parts[1])
                if (lon, lat) not in unique_coords:
                    unique_coords.append((lon, lat))
        if len(unique_coords) < 4:
            return None
        # Assuming order: top-left, top-right, bottom-right, bottom-left
        lon_min = unique_coords[0][0]
        lat_max = unique_coords[0][1]
        lon_max = unique_coords[1][0]
        lat_min = unique_coords[3][1]
        return lon_min, lon_max, lat_min, lat_max
    except Exception as e:
        logging.error(f"Error parsing KML file {kml_path}: {e}")
        return None

# ---------------------------------------------------------------------------
# Helpers for Excel processing
# ---------------------------------------------------------------------------
def set_calculation_properties(workbook):
    r"""Force auto calculation in openpyxl (if supported)."""
    try:
        workbook.calculation_properties.calculationMode = "auto"
        workbook.calculation_properties.calcCompleted = False
        workbook.calculation_properties.calcOnSave = True
        workbook.calculation_properties.fullCalcOnLoad = True
        workbook.calculation_properties.forceFullCalc = True
    except AttributeError:
        logging.warning("Calculation properties are not supported in this openpyxl version. Please upgrade.")

def force_recalc_with_excel(final_file):
    r"""Use Excel COM Automation to force full recalculation and save."""
    if not os.path.isfile(final_file):
        logging.error(f"Cannot recalc: File not found: {final_file}")
        return
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(final_file, UpdateLinks=False)
        wb.Application.CalculateFullRebuild()
        while wb.Application.CalculationState != 0:
            time.sleep(0.5)
        wb.Save()
        wb.Close(False)
        excel.Quit()
        logging.info(f"Excel recalculation forced and file saved: {final_file}")
    except Exception as e:
        logging.error(f"Error during Excel recalculation: {e}")

def is_indicator_complete(result_folder, indicator):
    r"""
    Consider the indicator complete if a chart file matching the pattern exists.
    The pattern used is: "chart_.*{indicator}.*\.xlsx"
    """
    pattern = r"chart_.*" + re.escape(indicator) + r".*\.xlsx"
    chart_file = find_file(result_folder, pattern)
    return chart_file is not None

# ---------------------------------------------------------------------------
# PART A: Multiprocessing code
# ---------------------------------------------------------------------------
def read_points_from_xlsm_openpyxl(file_path):
    r"""Reads columns A and B from the FIRST sheet in file_path, returns list of (x, y)."""
    points = []
    try:
        wb = load_workbook(file_path, data_only=True, keep_vba=True)
        first_sheet_name = wb.sheetnames[0]
        sheet = wb[first_sheet_name]
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                points.append((float(row[0]), float(row[1])))
    except Exception as e:
        logging.error(f"Error reading file {file_path}: {e}")
    finally:
        wb.close()
    return points

def read_rgb_values_at_point(img, point):
    x, y = point
    row = img.height - int(y) - 1
    col = int(x)
    if 0 <= row < img.height and 0 <= col < img.width:
        return img.getpixel((col, row))
    return None

def process_image(image_path, points):
    img = Image.open(image_path).convert("RGB")
    img_width, img_height = img.size
    scaled_points = [(x * img_width / 512, y * img_height / 512) for x, y in points]
    results = []
    for point in scaled_points:
        rgb = read_rgb_values_at_point(img, point)
        if rgb:
            r, g, b = rgb
            grayscale = 0.299 * r + 0.587 * g + 0.114 * b
            results.append([point[0], point[1],
                            round(r, 3), round(g, 3), round(b, 3),
                            round(grayscale, 3)])
    img.close()
    return results

def write_results_to_excel(results, workbook_path):
    wb = Workbook()
    wb.remove(wb.active)
    for index, data in enumerate(results):
        sheet = wb.create_sheet(title=str(index))
        sheet.append(['X', 'Y', 'R', 'G', 'B', 'Grayscale'])
        for row in data:
            sheet.append(row)
    set_calculation_properties(wb)
    wb.save(workbook_path)

def run_partA(input_grid_folder, input_images_folder, output_folder, indicator):
    r"""
    Look in input_grid_folder for a file matching a pattern like:
    "grid_.*{indicator}.*\.xlsm" and read from the first sheet.
    Then process all .jpg files in input_images_folder.
    """
    pattern = r"grid_.*" + re.escape(indicator) + r".*\.xlsm"
    grid_file = find_file(input_grid_folder, pattern)
    if not grid_file:
        logging.error(f"Grid file not found in {input_grid_folder} for indicator {indicator}")
        return False

    points = read_points_from_xlsm_openpyxl(grid_file)
    if not points:
        logging.error(f"No points read from {grid_file}. Exiting Part A.")
        return False

    jpg_files = [
        os.path.join(input_images_folder, f)
        for f in os.listdir(input_images_folder)
        if f.lower().endswith('.jpg')
    ]
    if not jpg_files:
        logging.error(f"No .jpg files found in {input_images_folder}.")
        return False

    tasks = [(jpg_path, points) for jpg_path in jpg_files]
    with Pool(cpu_count()) as pool:
        results = pool.starmap(process_image, tasks)

    output_xlsx = os.path.join(output_folder, f"results_{indicator}.xlsx")
    write_results_to_excel(results, output_xlsx)
    logging.info(f"Part A completed. Wrote {output_xlsx}")
    return True

# ---------------------------------------------------------------------------
# PART C: Process total results -> total_results_{indicator}.xlsx
# ---------------------------------------------------------------------------
def run_partC(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("~$"):
            continue
        if filename.endswith(".xlsx") and not filename.startswith("chart_"):
            input_file = os.path.join(output_folder, filename)
            match = re.search(r'_([A-Za-z0-9]+)\.xlsx$', filename)
            if not match:
                continue
            file_indicator = match.group(1)
            # Change file name here
            out_file = os.path.join(output_folder, f"total_results_{file_indicator}.xlsx")
            wb = load_workbook(input_file, read_only=True, data_only=True)
            data_frames = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = list(sheet.values)
                if not data:
                    continue
                data = data[1:]  # skip header
                df = pd.DataFrame(data)
                try:
                    chla_data = df.iloc[:, 3].dropna()
                    if not chla_data.empty:
                        data_frames.append(chla_data.reset_index(drop=True).to_frame(name=sheet_name))
                except IndexError:
                    logging.info(f"No Column 'D' in {sheet_name}")
            if data_frames:
                all_data_df = pd.concat(data_frames, axis=1)
                all_data_df['Average'] = all_data_df.mean(axis=1)
                with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                    all_data_df.to_excel(writer, sheet_name='ChlA', index=False)
                    wb2 = writer.book
                    set_calculation_properties(wb2)
                logging.info(f"Part C: Wrote {out_file}")
            else:
                logging.info(f"No data frames to concatenate for {filename}")

# ---------------------------------------------------------------------------
# PART D: Process charts -> chart_{indicator}.xlsx
# ---------------------------------------------------------------------------
def run_partD(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("~$"):
            continue
        if filename.startswith("total_results_") and filename.endswith(".xlsx"):
            source_file = os.path.join(output_folder, filename)
            # Force using the full indicator for naming (do not extract only digits)
            file_indicator = indicator  
            wb = load_workbook(source_file, data_only=True)
            if 'ChlA' not in wb.sheetnames:
                logging.info(f"Sheet 'ChlA' not found in {filename}")
                continue
            ws = wb['ChlA']
            avg_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == 'Average':
                    avg_col = col
                    break
            if avg_col is None:
                logging.info("Column 'Average' not found.")
                continue
            avg_data = [ws.cell(row=r, column=avg_col).value for r in range(2, ws.max_row + 1)]
            wb_out = Workbook()
            ws1 = wb_out.active
            ws1.title = 'Sheet1'
            for r_idx in range(1, 21):
                c = ws1.cell(row=r_idx, column=2, value=r_idx)
                c.number_format = "0.0000"
            col_index = 3
            for i in range(0, len(avg_data), 20):
                for j in range(20):
                    if (i + j) < len(avg_data):
                        c = ws1.cell(row=j+1, column=col_index, value=avg_data[i+j])
                        c.number_format = "0.0000"
                col_index += 1
            ws2 = wb_out.create_sheet(title='Sheet2')
            for col in range(3, col_index):
                col_letter = get_column_letter(col)
                for row in range(1, 21):
                    formula = f"=Sheet1!{col_letter}{row}/AVERAGE(Sheet1!{col_letter}$1:{col_letter}$20)"
                    ws2.cell(row=row, column=col, value=formula)
            dummy = wb_out.create_sheet(title='DummyCalc')
            dummy['A1'] = "=NOW()"
            dummy.sheet_state = 'hidden'
            set_calculation_properties(wb_out)
            intermediate_file = os.path.join(output_folder, f"chart_{file_indicator}_intermediate.xlsx")
            wb_out.save(intermediate_file)
            wb2 = load_workbook(intermediate_file)
            if 'Sheet2' in wb2.sheetnames:
                ws2_2 = wb2['Sheet2']
                ws3 = wb2.create_sheet(title='Sheet3')
                dest_row = 1
                for col in range(1, ws2_2.max_column + 1):
                    for row in range(1, ws2_2.max_row + 1):
                        val = ws2_2.cell(row=row, column=col).value
                        if val is not None:
                            ws3.cell(row=dest_row, column=2, value=val)
                            dest_row += 1
                set_calculation_properties(wb2)
                final_chart = os.path.join(output_folder, f"chart_{file_indicator}.xlsx")
                wb2.save(final_chart)
                logging.info(f"Part D: Chart file generated: {final_chart}")

# ---------------------------------------------------------------------------
# PART E: Export XLSM and CSV (Removed as per request)
# ---------------------------------------------------------------------------
def run_partE(output_folder, indicator):
    # Removed XLSM and CSV export as requested.
    logging.info(f"Part E: Skipping export of XLSM and CSV for indicator {indicator}")

# ---------------------------------------------------------------------------
# Reading final data for scatter from grid XLSM (first sheet)
# ---------------------------------------------------------------------------
def read_points_from_xlsm_pandas(file_path, sheet_name):
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    return df.iloc[:, :2].dropna().values

def read_data_from_xlsx(file_path, sheet_name, column_index):
    wb = load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in {file_path}")
    ws = wb[sheet_name]
    arr = []
    for row in ws.iter_rows(min_row=1, min_col=column_index+1, max_col=column_index+1, values_only=True):
        val = row[0]
        if val is not None:
            try:
                arr.append(float(val))
            except ValueError:
                pass
    return np.array(arr)

# ---------------------------------------------------------------------------
# SINGLE WINDOW + RADIO BUTTONS with Mirrored Axes
# ---------------------------------------------------------------------------
def run_interactive_scatter_for_all_indicators(all_data):
    r"""
    all_data: dict of
      indicator -> {
         'points': (pixel x_values, pixel y_values),
         'geo_points': (geo x_values, geo y_values) [optional],
         'geo_bounds': [lon_min, lon_max, lat_min, lat_max] [optional],
         'data': data array,
         'background': path to background image
      }
    Creates one figure with RadioButtons for indicator selection.
    If 'geo_points' and 'geo_bounds' are provided, these are used for plotting.
    """
    if not all_data:
        logging.error("No indicator data available for interactive scatter. Exiting.")
        return

    import matplotlib.pyplot as plt
    from PIL import Image

    indicators = list(all_data.keys())
    current_indicator = indicators[0]

    def get_current_data():
        return all_data[current_indicator]

    def update_scatter():
        d = get_current_data()
        # Use georeferenced points if available; otherwise, fallback to pixel points.
        if d.get('geo_points') is not None:
            xvals, yvals = d['geo_points']
            title_suffix = " (Geo)"
        else:
            xvals, yvals = d['points']
            title_suffix = " (Pixels)"
        arr   = d['data']
        img_path = d['background']
        new_img = Image.open(img_path)
        for im in ax1.images + ax2.images:
            im.set_data(new_img)
        new_sizes = np.full(len(arr), slider_size.val)
        new_colors = []
        for val in arr:
            dist = abs(val - slider_min.val) / max(slider_max.val - slider_min.val, 1e-6)
            gray = int(255 * np.exp(slider_exp.val * dist) * slider_sharp.val)
            gray = max(min(gray, 255), 0)
            new_colors.append((gray/255, gray/255, gray/255))
        scatter1.set_offsets(np.column_stack((xvals, yvals)))
        scatter1.set_sizes(new_sizes)
        scatter1.set_facecolors(new_colors)
        scatter1.set_alpha(slider_alpha1.val)
        scatter2.set_offsets(np.column_stack((xvals, yvals)))
        scatter2.set_sizes(new_sizes)
        scatter2.set_facecolors(new_colors)
        scatter2.set_alpha(slider_alpha2.val)
        ax1.set_title(f"Indicator: {current_indicator}{title_suffix} (Left Plot)")
        ax2.set_title(f"Indicator: {current_indicator}{title_suffix} (Right Plot)")
        fig.canvas.draw_idle()

    def on_indicator_label_clicked(label):
        nonlocal current_indicator
        current_indicator = label
        update_scatter()

    fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(14, 7))
    plt.subplots_adjust(left=0.25, bottom=0.35)

    updating_axes = [False]

    def on_xlim_changed(ax):
        if updating_axes[0]:
            return
        try:
            updating_axes[0] = True
            other_ax = ax2 if ax == ax1 else ax1
            other_ax.set_xlim(ax.get_xlim())
            fig.canvas.draw_idle()
        finally:
            updating_axes[0] = False

    def on_ylim_changed(ax):
        if updating_axes[0]:
            return
        try:
            updating_axes[0] = True
            other_ax = ax2 if ax == ax1 else ax1
            other_ax.set_ylim(ax.get_ylim())
            fig.canvas.draw_idle()
        finally:
            updating_axes[0] = False

    ax1.callbacks.connect('xlim_changed', on_xlim_changed)
    ax1.callbacks.connect('ylim_changed', on_ylim_changed)
    ax2.callbacks.connect('xlim_changed', on_xlim_changed)
    ax2.callbacks.connect('ylim_changed', on_ylim_changed)

    # Set the background image extent
    initial_data = get_current_data()
    if initial_data.get('geo_bounds'):
        # Use the geographic bounds from the KML.
        geo_bounds = initial_data['geo_bounds']
        # imshow expects extent in [left, right, bottom, top]
        image_extent = [geo_bounds[0], geo_bounds[1], geo_bounds[2], geo_bounds[3]]
    else:
        image_extent = [0, 512, 0, 512]
    
    background_img = Image.open(initial_data['background'])
    ax1.imshow(background_img, extent=image_extent, aspect='auto')
    scatter1 = ax1.scatter([], [], c='white')
    ax1.set_title(f"Indicator: {current_indicator} (Left Plot)")
    
    ax2.imshow(background_img, extent=image_extent, aspect='auto')
    scatter2 = ax2.scatter([], [], c='white')
    ax2.set_title(f"Indicator: {current_indicator} (Right Plot)")

    ax_size   = plt.axes([0.25, 0.2, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_min    = plt.axes([0.25, 0.17, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_max    = plt.axes([0.25, 0.14, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_alpha1 = plt.axes([0.25, 0.11, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_alpha2 = plt.axes([0.25, 0.08, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_exp    = plt.axes([0.25, 0.05, 0.65, 0.02], facecolor='lightgoldenrodyellow')
    ax_sharp  = plt.axes([0.25, 0.02, 0.65, 0.02], facecolor='lightgoldenrodyellow')

    slider_size   = Slider(ax_size,   'Size',     1, 30,   valinit=29.9)
    slider_min    = Slider(ax_min,    'Min',      0.0, 1.0, valinit=0.838)
    slider_max    = Slider(ax_max,    'Max',      1.0, 2.0, valinit=1.252)
    slider_alpha1 = Slider(ax_alpha1, 'Alpha1',   0.0, 1.0, valinit=0.687)
    slider_alpha2 = Slider(ax_alpha2, 'Alpha2',   0.0, 1.0, valinit=0.05)
    slider_exp    = Slider(ax_exp,    'Exponent', -10, 0,   valinit=-2.02)
    slider_sharp  = Slider(ax_sharp,  'Sharp',    0.1, 2.0, valinit=1.192)

    ax_radio = plt.axes([0.05, 0.35, 0.15, 0.5], facecolor='lightgoldenrodyellow')
    radio_buttons = RadioButtons(ax_radio, indicators, active=0)
    radio_buttons.on_clicked(on_indicator_label_clicked)

    def slider_updated(val):
        update_scatter()
    slider_size.on_changed(slider_updated)
    slider_min.on_changed(slider_updated)
    slider_max.on_changed(slider_updated)
    slider_alpha1.on_changed(slider_updated)
    slider_alpha2.on_changed(slider_updated)
    slider_exp.on_changed(slider_updated)
    slider_sharp.on_changed(slider_updated)

    update_scatter()
    plt.show()

# ---------------------------------------------------------------------------
# MAIN: Single-window UI with flexible naming and skip re-run if complete
# ---------------------------------------------------------------------------
def main_option2_single_window():
    r"""
    1) Scan for mother folders (e.g., "8", "pipe", "pipe_45", etc.) in working_dir.
    2) For each, check for subfolders: <indicator>_grid, <indicator>_jpgs, and <indicator>_result.
    3) If the result folder already contains a chart file, skip reprocessing.
    4) Collect scatter data from the first sheet of the grid XLSM and from the chart.
    5) If a KML file is present in the mother folder, use its coordinates to georeference the points and transform the background image.
    6) Launch one interactive UI for all indicators.
    """
    mother_folders = []
    for folder in os.listdir(working_dir):
        candidate = os.path.join(working_dir, folder)
        if os.path.isdir(candidate):
            grid_sub = os.path.join(candidate, f"{folder}_grid")
            jpgs_sub = os.path.join(candidate, f"{folder}_jpgs")
            result_sub = os.path.join(candidate, f"{folder}_result")
            if os.path.isdir(grid_sub) and os.path.isdir(jpgs_sub):
                if not os.path.isdir(result_sub):
                    os.makedirs(result_sub, exist_ok=True)
                mother_folders.append(folder)
    if not mother_folders:
        logging.error("No mother folders found (e.g., '8', 'pipe', 'pipe_45') with subfolders <indicator>_grid and <indicator>_jpgs.")
        return
    all_scatter_data = {}
    for indicator in mother_folders:
        mother_dir = os.path.join(working_dir, indicator)
        # Look for a KML file in the mother folder (if any)
        kml_file = find_file(mother_dir, r".*\.kml")
        kml_bounds = parse_kml_bounds(kml_file) if kml_file else None

        grid_dir = os.path.join(mother_dir, f"{indicator}_grid")
        jpgs_dir = os.path.join(mother_dir, f"{indicator}_jpgs")
        result_dir = os.path.join(mother_dir, f"{indicator}_result")
        if is_indicator_complete(result_dir, indicator):
            logging.info(f"Indicator '{indicator}' is already complete. Skipping re-run.")
        else:
            logging.info(f"Processing indicator: {indicator}")
            if not run_partA(grid_dir, jpgs_dir, result_dir, indicator):
                logging.error(f"Indicator {indicator}: Part A failed. Skipping.")
                continue
            run_partC(result_dir, indicator)
            run_partD(result_dir, indicator)
            run_partE(result_dir, indicator)
            chart_path = os.path.join(result_dir, f"chart_{indicator}.xlsx")
            force_recalc_with_excel(chart_path)
        grid_pattern = r"grid_.*" + re.escape(indicator) + r".*\.xlsm"
        chart_pattern = r"chart_.*" + re.escape(indicator) + r".*\.xlsx"
        grid_file = find_file(grid_dir, grid_pattern)
        chart_file = find_file(result_dir, chart_pattern)
        if not grid_file or not chart_file:
            logging.warning(f"Cannot load scatter data for {indicator}: missing grid or chart file.")
            continue
        points = read_points_from_xlsm_openpyxl(grid_file)
        if not points:
            logging.warning(f"No coordinate points found in {grid_file}. Skipping {indicator}.")
            continue
        x_vals = [p[0] for p in points]
        y_vals = [p[1] for p in points]
        # If KML bounds are available, compute georeferenced points and store geo_bounds.
        geo_points = None
        geo_bounds = None
        if kml_bounds:
            lon_min, lon_max, lat_min, lat_max = kml_bounds
            geo_bounds = [lon_min, lon_max, lat_min, lat_max]
            geo_x_vals = [lon_min + (x/512) * (lon_max - lon_min) for x in x_vals]
            # Adjusted vertical transformation (no mirroring)
            geo_y_vals = [lat_min + (y/512) * (lat_max - lat_min) for y in y_vals]
            geo_points = (geo_x_vals, geo_y_vals)
        try:
            wb = load_workbook(chart_file, data_only=True)
            if 'Sheet3' not in wb.sheetnames:
                logging.error(f"Sheet3 not found in {chart_file}.")
                continue
            ws = wb['Sheet3']
            data_array = []
            for row in ws.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True):
                val = row[0]
                if isinstance(val, (int, float)):
                    data_array.append(float(val))
            wb.close()
        except Exception as e:
            logging.error(f"Failed reading data from {chart_file}: {e}")
            continue
        frame0 = os.path.join(jpgs_dir, 'frame_0.jpg')
        if not os.path.isfile(frame0):
            cand = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
            if cand:
                frame0 = os.path.join(jpgs_dir, cand[0])
            else:
                logging.error(f"No background .jpg found in {jpgs_dir}. Skipping {indicator}.")
                continue
        all_scatter_data[indicator] = {
            'points': (x_vals, y_vals),
            'geo_points': geo_points,
            'geo_bounds': geo_bounds,
            'data': np.array(data_array),
            'background': frame0
        }
    run_interactive_scatter_for_all_indicators(all_scatter_data)

if __name__ == '__main__':
    main_option2_single_window()
