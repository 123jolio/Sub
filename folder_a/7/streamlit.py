import sys
import os
import re
import logging
import time
import numpy as np
import pandas as pd
import json
import tempfile, zipfile
from datetime import datetime
from multiprocessing import Pool, cpu_count, freeze_support
from PIL import Image, ImageOps
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use("Agg")  # Use a non-interactive backend for Streamlit
import matplotlib.pyplot as plt
from matplotlib.path import Path as MplPath
import xml.etree.ElementTree as ET
from pathlib import Path
from pyproj import Transformer
import win32com.client as win32
import contextily as ctx
import pytesseract

import streamlit as st

# Global variables
working_dir = os.getcwd()
ref_width = 512
ref_height = 512

# Configure logging (logs will be saved in app.log)
logging.basicConfig(filename="app.log", level=logging.INFO,
                    format="%(asctime)s - %(levelname)s - %(message)s")

#############################################
# HELPER FUNCTIONS
#############################################

def gif_to_frames(gif_path, output_folder):
    """Convert a GIF into individual JPEG frames saved in output_folder."""
    with Image.open(gif_path) as img:
        i = 0
        while True:
            try:
                img.seek(i)
                frame = img.convert('RGB')
                frame.save(os.path.join(output_folder, f'frame_{i}.jpg'), 'JPEG')
                i += 1
            except EOFError:
                break
    logging.info("GIF processed. Frames saved in: " + output_folder)

def find_file(folder, pattern):
    """Return the first file in folder matching the regex pattern (case-insensitive)."""
    regex = re.compile(pattern, re.IGNORECASE)
    for fname in os.listdir(folder):
        if regex.search(fname):
            return os.path.join(folder, fname)
    return None

def parse_kml_bounds(kml_path):
    """
    Extract (lon_min, lon_max, lat_min, lat_max) from a KML file
    by scanning all coordinates and computing min/max.
    """
    try:
        tree = ET.parse(kml_path)
        root = tree.getroot()
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        coords_text = root.find('.//kml:coordinates', ns).text.strip()
        coords = coords_text.split()
        all_lons, all_lats = [], []
        for coord in coords:
            parts = coord.split(',')
            if len(parts) >= 2:
                lon, lat = float(parts[0]), float(parts[1])
                all_lons.append(lon)
                all_lats.append(lat)
        if not all_lons:
            return None
        lon_min, lon_max = min(all_lons), max(all_lons)
        lat_min, lat_max = min(all_lats), max(all_lats)
        return (lon_min, lon_max, lat_min, lat_max)
    except Exception as e:
        logging.error(f"Error parsing KML file {kml_path}: {e}")
        return None

def parse_txt_bounds(txt_path):
    """
    Extract (lon_min, lon_max, lat_min, lat_max) from a TXT file containing JSON data.
    Expected format: {"type":"Polygon","coordinates":[[[lon,lat],[lon,lat], ...]]}
    Uses min/max of all coordinates.
    """
    try:
        with open(txt_path, 'r', encoding='utf-8-sig') as f:
            content = f.read().strip()
            if not content:
                logging.error(f"File {txt_path} is empty!")
                return None
            data = json.loads(content)
        if data.get("type") == "Polygon" and "coordinates" in data:
            coords = data["coordinates"][0]
            lons = [pt[0] for pt in coords]
            lats = [pt[1] for pt in coords]
            lon_min, lon_max = min(lons), max(lons)
            lat_min, lat_max = min(lats), max(lats)
            return (lon_min, lon_max, lat_min, lat_max)
        else:
            logging.error(f"JSON in {txt_path} does not have expected keys.")
            return None
    except Exception as e:
        logging.error(f"Error parsing txt file {txt_path}: {e}")
        return None

def set_calculation_properties(workbook):
    try:
        cp = workbook.calculation_properties
        cp.calculationMode = "auto"
        cp.calcCompleted = False
        cp.calcOnSave = True
        cp.fullCalcOnLoad = True
        cp.forceFullCalc = True
    except AttributeError:
        logging.warning("Calculation properties not supported in this openpyxl version.")

def force_recalc_with_excel(final_file):
    if not os.path.isfile(final_file):
        logging.error(f"Cannot recalc: File not found: {final_file}")
        return
    try:
        excel = win32.Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(final_file, UpdateLinks=False)
        wb.Application.CalculateFullRebuild()
        while wb.Application.CalculationState != 0:
            time.sleep(0.5)
        wb.Save()
        wb.Close(False)
        excel.Quit()
        logging.info(f"Excel recalculation forced and file saved: {final_file}")
    except Exception as e:
        logging.error(f"Error during Excel recalculation: {e}")

def is_indicator_complete(result_folder, indicator):
    pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
    return bool(find_file(result_folder, pattern))

def read_rgb_values_at_point(image, point):
    try:
        pixel = image.getpixel((int(point[0]), int(point[1])))
        if isinstance(pixel, (int, float)):
            return (pixel, pixel, pixel)
        return pixel[:3]
    except Exception as e:
        logging.error(f"Error reading pixel at {point}: {e}")
        return None

def read_points_from_xlsx_openpyxl(file_path):
    points = []
    grid_shape = (None, None)
    try:
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[wb.sheetnames[0]]
        points = [(float(row[0]), float(row[1])) for row in sheet.iter_rows(min_row=2, values_only=True)
                  if row[0] is not None and row[1] is not None]
        n_rows = sheet['D2'].value
        n_cols = sheet['E2'].value
        grid_shape = (int(n_rows) if n_rows is not None else None,
                      int(n_cols) if n_cols is not None else None)
    except Exception as e:
        logging.error(f"Error reading grid file {file_path}: {e}")
    return points, grid_shape

def get_time_series_for_points(results_file, indices):
    wb = load_workbook(results_file, data_only=True)
    time_series = {idx: [] for idx in indices}
    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx in indices:
            try:
                val = float(rows[idx][5])
            except (IndexError, TypeError):
                val = None
            time_series[idx].append(val)
    wb.close()
    return time_series

# -------------------------
# OCR FUNCTIONS
# -------------------------
pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"

def pil_multi_threshold_preprocess(roi, thresholds=[80,120,160,200], invert=True):
    gray = roi.convert("L")
    for thr in thresholds:
        bin_img = gray.point(lambda p: 255 if p > thr else 0)
        yield bin_img
        if invert:
            yield ImageOps.invert(bin_img)

def pil_attempt_ocr(preprocessed_images, config):
    for img in preprocessed_images:
        text = pytesseract.image_to_string(img, config=config).strip()
        found = re.findall(r"\d{4}-\d{2}-\d{2}", text)
        if found:
            dt_str = found[0]
            try:
                return datetime.strptime(dt_str, "%Y-%m-%d")
            except:
                continue
    return None

def extract_date(image_path):
    config = "--psm 7 -c tessedit_char_whitelist=0123456789-"
    try:
        img = Image.open(str(image_path))
    except Exception as e:
        logging.error(f"Error: Unable to load image at {image_path}: {e}")
        return None
    width, height = img.size
    roi_dyn = img.crop((int(0.75*width), 0, width, int(0.15*height)))
    roi_dyn = roi_dyn.resize((roi_dyn.width*2, roi_dyn.height*2), Image.LANCZOS)
    dyn_images = pil_multi_threshold_preprocess(roi_dyn, thresholds=[80,120,160,200], invert=True)
    date_dyn = pil_attempt_ocr(dyn_images, config)
    if date_dyn:
        logging.info(f"Dynamic OCR from {Path(image_path).name}: recognized {date_dyn.strftime('%Y-%m-%d')}")
        return date_dyn
    else:
        logging.info(f"Dynamic OCR failed for {Path(image_path).name}, attempting fixed cropping.")
    roi_fixed = img.crop((width-160, 0, width, 60))
    roi_fixed = roi_fixed.resize((roi_fixed.width*2, roi_fixed.height*2), Image.LANCZOS)
    fixed_images = pil_multi_threshold_preprocess(roi_fixed, thresholds=[80,120,160,200], invert=True)
    date_fixed = pil_attempt_ocr(fixed_images, config)
    if date_fixed:
        logging.info(f"Fixed OCR from {Path(image_path).name}: recognized {date_fixed.strftime('%Y-%m-%d')}")
        return date_fixed
    else:
        logging.info(f"Date not found in {Path(image_path).name}")
        return None

def extract_number(filename):
    number = re.findall(r'\d+$', Path(filename).stem)
    return int(number[0]) if number else 0

def save_ocr_results(jpgs_dir):
    folder_path = Path(jpgs_dir)
    data = []
    for filename in sorted(folder_path.iterdir()):
        if filename.suffix.lower() in ('.png','.jpg','.jpeg'):
            frame_num = extract_number(filename)
            dt = extract_date(filename)
            data.append((frame_num, dt))
    data.sort(key=lambda x: x[0])
    wb_ocr = Workbook()
    ws = wb_ocr.active
    ws.title = "OCR_Dates"
    ws.append(["FrameNumber", "RecognizedDate"])
    for (fnum, dt) in data:
        dt_str = dt.strftime("%Y-%m-%d") if dt else ""
        ws.append([fnum, dt_str])
    out_path = folder_path / "ocr_results.xlsx"
    wb_ocr.save(str(out_path))
    logging.info(f"OCR results saved to {out_path}")
    return out_path

def read_ocr_dates(jpgs_dir):
    folder_path = Path(jpgs_dir)
    ocr_file = folder_path / "ocr_results.xlsx"
    if not ocr_file.exists():
        ocr_file = save_ocr_results(jpgs_dir)
    wb = load_workbook(str(ocr_file), data_only=True)
    ws = wb.active
    frame_dates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        frame_num = row[0]
        date_str = row[1]
        dt = None
        if date_str:
            try:
                dt = datetime.strptime(date_str, "%Y-%m-%d")
            except:
                pass
        frame_dates[frame_num] = dt
    wb.close()
    if frame_dates:
        max_frame = max(frame_dates.keys())
        out = [None]*(max_frame+1)
        for i in range(max_frame+1):
            out[i] = frame_dates.get(i, None)
        return out
    else:
        return []

#############################################
# GRID CREATION FUNCTION (Streamlit version)
#############################################

def create_grid_streamlit(indicator, jpgs_dir):
    """
    Grid creation using Streamlit widgets.
    Two modes: Polygon mode (enter vertices manually) and Origin mode.
    """
    st.subheader("Grid Creation")
    mode = st.radio("Select grid creation mode:", ("Polygon", "Origin"))
    # Load a background image from jpgs_dir
    candidates = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
    if not candidates:
        st.error("No background image found in the specified directory.")
        return None
    bg_path = os.path.join(jpgs_dir, "frame_0.jpg") if os.path.isfile(os.path.join(jpgs_dir, "frame_0.jpg")) else os.path.join(jpgs_dir, candidates[0])
    try:
        bg_img = Image.open(bg_path)
    except Exception as e:
        st.error(f"Error opening background image: {e}")
        return None

    grid_pts = []
    if mode == "Polygon":
        st.info("Enter polygon vertices (x,y) one per line (e.g., `100,150`):")
        poly_input = st.text_area("Polygon Vertices", height=150)
        if poly_input:
            try:
                vertices = []
                for line in poly_input.strip().splitlines():
                    parts = line.split(',')
                    if len(parts) == 2:
                        vertices.append((float(parts[0].strip()), float(parts[1].strip())))
                if len(vertices) < 3:
                    st.error("A polygon must have at least 3 points.")
                    return None
            except Exception as e:
                st.error(f"Error parsing vertices: {e}")
                return None

            spacing_h_init = st.number_input("Initial horizontal grid spacing", value=50.0, min_value=1.0)
            spacing_v_init = spacing_h_init
            grid_rot_init = st.slider("Grid Rotation", -180.0, 180.0, 0.0)
            poly_rot_init = st.slider("Polygon Rotation", -180.0, 180.0, 0.0)

            def rotate_points(points, angle_deg, center):
                angle_rad = np.radians(angle_deg)
                cos_a, sin_a = np.cos(angle_rad), np.sin(angle_rad)
                return [[(pt[0]-center[0])*cos_a - (pt[1]-center[1])*sin_a + center[0],
                         (pt[0]-center[0])*sin_a + (pt[1]-center[1])*cos_a + center[1]] for pt in points]
            
            poly_arr = np.array(vertices)
            centroid = poly_arr.mean(axis=0)
            
            def generate_grid(h_spacing, v_spacing, poly_rot, grid_rot):
                adjusted_poly = rotate_points(vertices, poly_rot, centroid)
                poly_path = MplPath(adjusted_poly)
                xs, ys = zip(*adjusted_poly)
                bx_min, bx_max = min(xs), max(xs)
                by_min, by_max = min(ys), max(ys)
                pts = []
                for yy in np.arange(by_min, by_max + v_spacing, v_spacing):
                    for xx in np.arange(bx_min, bx_max + h_spacing, h_spacing):
                        pts.append([xx, yy])
                grid_center = [(bx_min+bx_max)/2, (by_min+by_max)/2]
                pts_rot = rotate_points(pts, grid_rot, grid_center)
                valid_pts = [pt for pt in pts_rot if poly_path.contains_point((pt[0], pt[1]))]
                valid_pts.sort(key=lambda pt: (pt[1], pt[0]))
                return adjusted_poly, valid_pts

            if st.button("Generate Grid (Polygon Mode)"):
                _, grid_pts = generate_grid(spacing_h_init, spacing_v_init, poly_rot_init, grid_rot_init)
                st.success(f"Generated {len(grid_pts)} grid points.")
                # Preview the grid
                fig, ax = plt.subplots()
                # We'll show it with origin='lower' for a standard bottom-up approach
                ax.imshow(bg_img, extent=[0, ref_width, 0, ref_height], origin='lower')
                if grid_pts:
                    xs, ys = zip(*grid_pts)
                    ax.scatter(xs, ys, color='red', s=10)
                poly = np.array(vertices)
                ax.plot(np.append(poly[:,0], poly[0,0]),
                        np.append(poly[:,1], poly[0,1]), color='blue', linestyle='--')
                st.pyplot(fig)
    else:  # Origin mode
        st.info("Enter origin (bottom-left) coordinates:")
        origin_x = st.number_input("Origin X", value=0.0)
        origin_y = st.number_input("Origin Y", value=0.0)
        init_spacing = st.number_input("Grid Spacing", value=50.0, min_value=1.0)
        init_rows = st.number_input("Number of Rows", value=10, min_value=1, step=1)
        init_cols = st.number_input("Number of Columns", value=10, min_value=1, step=1)
        init_angle = st.slider("Grid Angle", -90.0, 90.0, 0.0)
        
        if st.button("Generate Grid (Origin Mode)"):
            grid_pts = []
            import math
            angle_rad = math.radians(init_angle)
            for i in range(int(init_rows)):
                for j in range(int(init_cols)):
                    xx = j * init_spacing
                    yy = i * init_spacing
                    x_rot = xx * math.cos(angle_rad) - yy * math.sin(angle_rad)
                    y_rot = xx * math.sin(angle_rad) + yy * math.cos(angle_rad)
                    grid_pts.append([origin_x + x_rot, origin_y + y_rot])
            grid_pts.sort(key=lambda pt: (pt[1], pt[0]))
            st.success(f"Generated grid with {len(grid_pts)} points.")
            fig, ax = plt.subplots()
            ax.imshow(bg_img, extent=[0, ref_width, 0, ref_height], origin='lower')
            if grid_pts:
                xs, ys = zip(*grid_pts)
                ax.scatter(xs, ys, color='red', s=10)
            st.pyplot(fig)
            # Save grid shape from inputs
            rows = int(init_rows)
            cols = int(init_cols)

    if grid_pts:
        # Save grid points to Excel file
        wb = Workbook()
        ws = wb.active
        ws.title = "GridPoints"
        ws.append(["X", "Y"])
        for pt in grid_pts:
            ws.append(pt)
        if mode == "Polygon":
            ys = [pt[1] for pt in grid_pts]
            unique_y = np.unique(np.round(ys, 2))
            n_rows = len(unique_y)
            n_cols = int(np.ceil(len(grid_pts) / n_rows))
        else:
            n_rows = int(init_rows)
            n_cols = int(init_cols)
        ws['D1'] = "Rows"
        ws['D2'] = n_rows
        ws['E1'] = "Cols"
        ws['E2'] = n_cols
        grid_filename = f"grid_{indicator}.xlsx"
        save_path = os.path.join(working_dir, grid_filename)
        wb.save(save_path)
        logging.info(f"Grid file created: {save_path}")
        st.success(f"Grid file created: {save_path}")
        return save_path
    else:
        st.error("Grid generation failed.")
        return None

#############################################
# IMAGE PROCESSING & RESULTS FUNCTIONS
#############################################

def process_image(image_path, points):
    """
    Scale image points and extract RGB values, assuming bottom-left origin in local space.
    """
    img = Image.open(image_path).convert("RGB")
    img_width, img_height = img.size
    # If your local images are top-left origin, you might invert Y:
    # y_inverted = (ref_height - yy)
    # But we'll assume bottom-left for now:
    scaled_points = [
        (xx * img_width / ref_width,
         yy * img_height / ref_height)
        for xx, yy in points
    ]
    results = []
    for pt in scaled_points:
        rgb = read_rgb_values_at_point(img, pt)
        if rgb:
            r, g, b = rgb
            grayscale = 0.299*r + 0.587*g + 0.114*b
            results.append([pt[0], pt[1], round(r, 3), round(g, 3), round(b, 3), round(grayscale, 3)])
    img.close()
    return results

def write_results_to_excel(results, workbook_path):
    wb = Workbook()
    wb.remove(wb.active)
    for index, data in enumerate(results):
        sheet = wb.create_sheet(title=str(index))
        sheet.append(["X", "Y", "R", "G", "B", "Grayscale"])
        for row in data:
            sheet.append(row)
    set_calculation_properties(wb)
    wb.save(workbook_path)
    logging.info(f"Results written to {workbook_path}")

def run_partA(input_grid_folder, input_images_folder, output_folder, indicator):
    pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
    grid_file = find_file(input_grid_folder, pattern)
    if not grid_file:
        logging.error(f"Grid file not found in {input_grid_folder} for indicator {indicator}")
        return False
    points, _ = read_points_from_xlsx_openpyxl(grid_file)
    if not points:
        logging.error(f"No points read from {grid_file}. Exiting Part A.")
        return False
    jpg_files = [
        os.path.join(input_images_folder, f)
        for f in os.listdir(input_images_folder)
        if f.lower().endswith('.jpg')
    ]
    if not jpg_files:
        logging.error(f"No .jpg files found in {input_images_folder}.")
        return False
    with Pool(cpu_count()) as pool:
        results = pool.starmap(process_image, [(jpg, points) for jpg in jpg_files])
    output_xlsx = os.path.join(output_folder, f"results_{indicator}.xlsx")
    write_results_to_excel(results, output_xlsx)
    logging.info(f"Part A completed. Wrote {output_xlsx}")
    return True

def run_partC(output_folder, indicator):
    import pandas as pd
    for filename in os.listdir(output_folder):
        if filename.startswith("~$") or filename.startswith("chart_"):
            continue
        if filename.endswith(".xlsx"):
            input_file = os.path.join(output_folder, filename)
            match = re.search(r'_([A-Za-z0-9]+)\.xlsx$', filename)
            if not match:
                continue
            file_indicator = match.group(1)
            out_file = os.path.join(output_folder, f"total_results_{file_indicator}.xlsx")
            wb = load_workbook(input_file, read_only=True, data_only=True)
            data_frames = []
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                data = list(sheet.values)
                if not data:
                    continue
                data = data[1:]
                df = pd.DataFrame(data)
                try:
                    col_data = df.iloc[:, 3].dropna()
                    if not col_data.empty:
                        data_frames.append(col_data.reset_index(drop=True).to_frame(name=sheet_name))
                except IndexError:
                    logging.info(f"No Column 'D' in {sheet_name}")
            if data_frames:
                all_data_df = pd.concat(data_frames, axis=1)
                all_data_df['Average'] = all_data_df.mean(axis=1)
                with pd.ExcelWriter(out_file, engine='openpyxl') as writer:
                    all_data_df.to_excel(writer, sheet_name='ChlA', index=False)
                    set_calculation_properties(writer.book)
                logging.info(f"Part C: Wrote {out_file}")

def run_partD(output_folder, indicator):
    for filename in os.listdir(output_folder):
        if filename.startswith("~$"):
            continue
        if filename.startswith("total_results_") and filename.endswith(".xlsx"):
            source_file = os.path.join(output_folder, filename)
            file_indicator = indicator
            wb = load_workbook(source_file, data_only=True)
            if 'ChlA' not in wb.sheetnames:
                logging.info(f"Sheet 'ChlA' not found in {filename}.")
                continue
            ws = wb['ChlA']
            avg_col = None
            for col in range(1, ws.max_column+1):
                if ws.cell(row=1, column=col).value == 'Average':
                    avg_col = col
                    break
            if avg_col is None:
                logging.info("Column 'Average' not found.")
                continue
            avg_data = [ws.cell(row=r, column=avg_col).value for r in range(2, ws.max_row+1)]
            grid_file = find_file(os.path.join(os.path.dirname(source_file).replace("_result", "_grid")),
                                   rf"grid_.*{re.escape(indicator)}.*\.xlsx")
            _, grid_shape = read_points_from_xlsx_openpyxl(grid_file) if grid_file else (None, (20,))
            grid_rows = grid_shape[0] if grid_shape[0] is not None else 20
            wb_out = Workbook()
            ws1 = wb_out.active
            ws1.title = 'Sheet1'
            for r in range(1, grid_rows+1):
                ws1.cell(row=r, column=2, value=r).number_format = "0.0000"
            col_index = 3
            for i in range(0, len(avg_data), grid_rows):
                for j in range(grid_rows):
                    if i+j < len(avg_data):
                        ws1.cell(row=j+1, column=col_index, value=avg_data[i+j]).number_format = "0.0000"
                col_index += 1
            ws2 = wb_out.create_sheet(title='Sheet2')
            for col in range(3, col_index):
                col_letter = get_column_letter(col)
                for row in range(1, grid_rows+1):
                    ws2.cell(row=row, column=col, value=(
                        f"=Sheet1!{col_letter}{row}/AVERAGE(Sheet1!{col_letter}$1:{col_letter}${grid_rows})"
                    ))
            dummy = wb_out.create_sheet(title='DummyCalc')
            dummy['A1'] = "=NOW()"
            dummy.sheet_state = 'hidden'
            set_calculation_properties(wb_out)
            intermediate_file = os.path.join(output_folder, f"chart_{file_indicator}_intermediate.xlsx")
            wb_out.save(intermediate_file)
            wb2 = load_workbook(intermediate_file)
            if 'Sheet2' in wb2.sheetnames:
                ws2_2 = wb2['Sheet2']
                ws3 = wb2.create_sheet(title='Sheet3')
                dest_row = 1
                for col in range(1, ws2_2.max_column+1):
                    for row in range(1, ws2_2.max_row+1):
                        val = ws2_2.cell(row=row, column=col).value
                        if val is not None:
                            ws3.cell(row=dest_row, column=2, value=val)
                            dest_row += 1
                set_calculation_properties(wb2)
                final_chart = os.path.join(output_folder, f"chart_{file_indicator}.xlsx")
                wb2.save(final_chart)
                logging.info(f"Part D: Chart file generated: {final_chart}")

def run_partE(output_folder, indicator):
    logging.info(f"Part E: Skipping export of XLSM and CSV for indicator {indicator}")

#############################################
# INTERACTIVE SCATTER & EVOLUTION (Updated)
#############################################

def run_interactive_scatter_for_all_indicators(all_data):
    """
    Plots each indicator on a static scatter with improved bounding box handling
    and a visible scatter color (red).
    """
    if not all_data:
        logging.error("No indicator data available for interactive scatter. Exiting.")
        return
    
    for indicator, data_dict in all_data.items():
        st.subheader(f"Scatter Plot for Indicator: {indicator}")
        
        # If geo_points is set, we have lat/lon that was turned into geo_points
        # We'll transform them to EPSG:3857 for a consistent map projection
        if data_dict.get('geo_points'):
            lon_min, lon_max, lat_min, lat_max = data_dict['geo_bounds']
            transformer = Transformer.from_crs("EPSG:4326", "EPSG:3857", always_xy=True)
            xvals, yvals = data_dict['geo_points']  # these are in lat/lon, presumably
            xvals_trans = []
            yvals_trans = []
            for lon, lat in zip(xvals, yvals):
                x_t, y_t = transformer.transform(lon, lat)
                xvals_trans.append(x_t)
                yvals_trans.append(y_t)
            
            # Transform bounding box corners
            x_min, y_min = transformer.transform(lon_min, lat_min)
            x_max, y_max = transformer.transform(lon_max, lat_max)
            
            default_extent = [x_min, x_max, y_min, y_max]
            
            try:
                bg_img = Image.open(data_dict['background'])
            except Exception as e:
                st.error(f"Error opening background image: {e}")
                continue
            
            fig, ax = plt.subplots(figsize=(7, 5))
            # Standard bottom-up approach
            ax.imshow(bg_img, extent=default_extent, origin='lower', aspect='auto')
            ax.scatter(xvals_trans, yvals_trans, color='red', s=50)
            ax.set_xlim(x_min, x_max)
            ax.set_ylim(y_min, y_max)
            ax.set_title(f"Indicator: {indicator}")
            st.pyplot(fig)
        
        # Otherwise, local (non-georeferenced) approach
        else:
            try:
                bg_img = Image.open(data_dict['background'])
                w, h = bg_img.size
            except Exception as e:
                st.error(f"Error opening background image: {e}")
                w, h = ref_width, ref_height
            
            xvals, yvals = data_dict['points']
            # If your local approach is bottom-left, we do:
            xvals_scaled = np.array(xvals) * (w / ref_width)
            yvals_scaled = np.array(yvals) * (h / ref_height)
            
            default_extent = [0, w, 0, h]
            
            fig, ax = plt.subplots(figsize=(7, 5))
            ax.imshow(bg_img, extent=default_extent, origin='lower', aspect='auto')
            ax.scatter(xvals_scaled, yvals_scaled, color='red', s=50)
            ax.set_title(f"Indicator: {indicator}")
            st.pyplot(fig)

#############################################
# MAIN PROCESSING FUNCTION (Streamlit version)
#############################################

def main_option2_streamlit():
    st.header("Processing Indicators")
    # Look for mother folders in working_dir that contain required subfolders
    mother_folders = [folder for folder in os.listdir(working_dir)
                      if os.path.isdir(os.path.join(working_dir, folder))]
    valid_folders = []
    for folder in mother_folders:
        candidate = os.path.join(working_dir, folder)
        grid_sub = os.path.join(candidate, f"{folder}_grid")
        jpgs_sub = os.path.join(candidate, f"{folder}_jpgs")
        result_sub = os.path.join(candidate, f"{folder}_result")
        if os.path.isdir(grid_sub) and os.path.isdir(jpgs_sub):
            os.makedirs(result_sub, exist_ok=True)
            valid_folders.append(folder)
    if not valid_folders:
        st.error("No valid mother folders found with required subfolders.")
        logging.error("No valid mother folders found with required subfolders.")
        return

    indicator = st.selectbox("Select Indicator", valid_folders)
    mother_dir = os.path.join(working_dir, indicator)
    grid_dir = os.path.join(mother_dir, f"{indicator}_grid")
    jpgs_dir = os.path.join(mother_dir, f"{indicator}_jpgs")
    result_dir = os.path.join(mother_dir, f"{indicator}_result")

    # If no JPG files, try converting a GIF
    jpgs = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
    if not jpgs:
        gif_files = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.gif')]
        if gif_files:
            gif_path = os.path.join(jpgs_dir, gif_files[0])
            st.info(f"No JPG found in {jpgs_dir}. Converting GIF: {gif_path}")
            gif_to_frames(gif_path, jpgs_dir)

    # If no grid file exists, offer grid creation
    if not find_file(grid_dir, r"grid_.*\.xlsx"):
        st.warning(f"No grid file found for indicator '{indicator}'.")
        if st.button("Create Grid"):
            grid_path = create_grid_streamlit(indicator, jpgs_dir)
            if grid_path:
                os.makedirs(grid_dir, exist_ok=True)
                new_path = os.path.join(grid_dir, os.path.basename(grid_path))
                os.replace(grid_path, new_path)
                st.success(f"Grid created and moved to {new_path}")
            else:
                st.error(f"Grid creation failed for indicator '{indicator}'.")
                return

    # Process indicator if not already complete
    if is_indicator_complete(result_dir, indicator):
        st.info(f"Indicator '{indicator}' is already complete. Skipping re-run.")
    else:
        st.info(f"Processing indicator: {indicator}")
        if not run_partA(grid_dir, jpgs_dir, result_dir, indicator):
            st.error(f"Indicator {indicator}: Part A failed. Skipping.")
            return
        run_partC(result_dir, indicator)
        run_partD(result_dir, indicator)
        run_partE(result_dir, indicator)
        chart_path = os.path.join(result_dir, f"chart_{indicator}.xlsx")
        force_recalc_with_excel(chart_path)

    grid_pattern = rf"grid_.*{re.escape(indicator)}.*\.xlsx"
    chart_pattern = rf"chart_.*{re.escape(indicator)}.*\.xlsx"
    grid_file = find_file(grid_dir, grid_pattern)
    chart_file = find_file(result_dir, chart_pattern)
    if not grid_file or not chart_file:
        st.warning(f"Missing grid or chart file for indicator '{indicator}'.")
        return

    points, _ = read_points_from_xlsx_openpyxl(grid_file)
    if not points:
        st.warning(f"No points in grid file for indicator '{indicator}'. Skipping.")
        return

    x_vals, y_vals = zip(*points)
    bounds = None
    kml_file = find_file(mother_dir, r".*\.kml")
    if kml_file:
        bounds = parse_kml_bounds(kml_file)
    else:
        txt_file = find_file(mother_dir, r".*\.txt")
        if txt_file:
            bounds = parse_txt_bounds(txt_file)

    geo_points = None
    geo_bounds = None
    if bounds:
        lon_min, lon_max, lat_min, lat_max = bounds
        geo_bounds = [lon_min, lon_max, lat_min, lat_max]
        # If we interpret (0,0) as bottom-left for local image coords:
        # x=0 => lon_min, x=ref_width => lon_max
        # y=0 => lat_min, y=ref_height => lat_max
        # But if you used top-left references, invert Y as needed
        geo_x_vals = [lon_min + (xx / ref_width) * (lon_max - lon_min) for xx in x_vals]
        geo_y_vals = [lat_min + (yy / ref_height) * (lat_max - lat_min) for yy in y_vals]
        geo_points = (geo_x_vals, geo_y_vals)

    try:
        chart_file_path = os.path.join(result_dir, f"chart_{indicator}.xlsx")
        wb = load_workbook(chart_file_path, data_only=True)
        if 'Sheet3' not in wb.sheetnames:
            st.error(f"Sheet3 missing in {chart_file_path}.")
            logging.error(f"Sheet3 missing in {chart_file_path}.")
            return
        ws = wb['Sheet3']
        data_array = [float(row[0]) for row in ws.iter_rows(min_row=1, min_col=2, max_col=2, values_only=True)
                      if isinstance(row[0], (int, float))]
        wb.close()
    except Exception as e:
        st.error(f"Error reading data from {chart_file_path}: {e}")
        logging.error(f"Error reading data from {chart_file_path}: {e}")
        return

    frame0 = os.path.join(jpgs_dir, 'frame_0.jpg')
    if not os.path.isfile(frame0):
        cand = [f for f in os.listdir(jpgs_dir) if f.lower().endswith('.jpg')]
        if cand:
            frame0 = os.path.join(jpgs_dir, cand[0])
        else:
            st.error(f"No background image in {jpgs_dir}. Skipping {indicator}.")
            logging.error(f"No background image in {jpgs_dir}. Skipping {indicator}.")
            return

    results_file = os.path.join(result_dir, f"results_{indicator}.xlsx")
    all_data = {}
    all_data[indicator] = {
        'points': (list(x_vals), list(y_vals)),
        'geo_points': geo_points,
        'geo_bounds': geo_bounds,
        'data': np.array(data_array),
        'background': frame0,
        'results_file': results_file
    }
    run_interactive_scatter_for_all_indicators(all_data)

#############################################
# MAIN STREAMLIT APP STRUCTURE
#############################################

def main():
    st.title("Streamlit Data Processing Application")
    st.write("This application processes images and grids to generate results and interactive plots.")
    
    menu = st.sidebar.radio("Navigation", ["Home", "Process", "Logs"])
    
    if menu == "Home":
        st.write("Welcome to the application. Use the sidebar to navigate.")
    elif menu == "Process":
        main_option2_streamlit()
    elif menu == "Logs":
        st.subheader("Application Logs")
        if st.button("Refresh Logs"):
            try:
                with open("app.log", "r") as f:
                    log_text = f.read()
            except FileNotFoundError:
                log_text = "No log file found."
            st.text_area("Logs output", log_text, height=300)

if __name__ == '__main__':
    freeze_support()
    main()
